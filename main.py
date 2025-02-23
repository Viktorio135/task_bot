import os
import asyncio
import random
import datetime
import threading
import openpyxl
import shutil
import dill as pickle

from apscheduler.schedulers.asyncio import AsyncIOScheduler

from aiogram import Bot, Dispatcher, executor, types
from aiogram.dispatcher import FSMContext
from aiogram.dispatcher.filters import Text
from aiogram.contrib.fsm_storage.memory import MemoryStorage
from aiogram.utils.deep_linking import get_start_link, decode_payload
from aiogram.types import ContentType


from dotenv import load_dotenv
from functools import wraps

from database.db_commands import *
from keyboards.user_keyboards import *
from states.user_states import *
from database.models import start_db

load_dotenv()
user_token = os.getenv('USER_TOKEN')

user_bot = Bot(token=user_token, parse_mode='HTML', disable_web_page_preview=True)

user_dp = Dispatcher(user_bot, storage=MemoryStorage())

cached_data = {}
main_menu_icon = {}
tasks_progress = {}
archive_tasks_progerss = {}
warnings = {}
block_users = set()


## Проверка на блокировку
def block_check(func):
    @wraps(func)
    async def wrapped(message: types.Message, *args, **kwargs):
        user_id = message.from_user.id
        if user_id in block_users:
            return
        return await func(message, *args, **kwargs)
    return wrapped
##

## Проверка на админа 

channel_base_access = os.getenv('CHANNEL_BASE_ACCESS')
channel_vip_access = os.getenv('CHANNEL_VIP_ACCESS')

def allow_bace_access(func):
    @wraps(func)
    async def wrapped(message: types.Message, *args, **kwargs):        
        # Получаем информацию о членстве пользователя в канале
        base_access = await user_bot.get_chat_member(
            chat_id=channel_base_access,
            user_id=message.from_user.id
            )
        if base_access.status == 'left':
            return
        
        # Если все проверки пройдены, выполняем функцию
        return await func(message, *args, **kwargs)
    return wrapped

def allow_vip_access(func):
    @wraps(func)
    async def wrapped(message: types.Message, *args, **kwargs):        
        # Получаем информацию о членстве пользователя в канале
        base_access = await user_bot.get_chat_member(
            chat_id=channel_vip_access,
            user_id=message.from_user.id
            )
        if base_access.status == 'left':
            return
        
        # Если все проверки пройдены, выполняем функцию
        return await func(message, *args, **kwargs)
    return wrapped
##





############################################### User bot ##################################################

####### регистрация
@user_dp.message_handler(commands='start', state='*')
@block_check
async def uses_cmd_start(msg: types.Message, state: FSMContext):
    current_state = await state.get_state()
    if current_state is not None:  # Проверяем, находится ли пользователь в каком-либо состоянии
        await state.finish()  # Завершаем текущее состояние
    if not(await has_register(str(msg.from_user.id))):
        args = msg.get_args()
        if args != '':
            ref_code = decode_payload(args)
            cached_data[str(msg.from_user.id)] = ref_code
        user_channel_status = await user_bot.get_chat_member(
                chat_id='-1002171208182',
                user_id=msg.from_user.id
                )
        if user_channel_status.status != 'left':
            await user_bot.send_message(
                msg.from_user.id,
                'Вы должны быть подписаны на @asdasdsdfsv, чтобы пользоваться этим ботом.',
                reply_markup=check_sub()
            )
        else:
            await user_bot.send_message(
                msg.from_user.id,
                'Вы должны быть подписаны на @asdasdsdfsv, чтобы пользоваться этим ботом.',
                reply_markup=check_sub()
            )
    else:
        await change_is_admin(str(msg.from_user.id), is_admin=False)
        await user_main_menu(msg, has_register=True)

@user_dp.callback_query_handler(lambda c: c.data == 'check_sub')
@block_check
async def user_check_sub_start(callback_query: types.CallbackQuery, start=True):
    user_channel_status = await user_bot.get_chat_member(
            chat_id='-1002171208182',
            user_id=callback_query.from_user.id
            )
    if user_channel_status.status != 'left':
        ref_code = ''
        if str(callback_query.from_user.id) in cached_data:
            ref_code = cached_data[str(callback_query.from_user.id)]
            await adding_referal(ref_code)
            del cached_data[str(callback_query.from_user.id)]
        if start:
            await create_user(
                user_name=callback_query.from_user.username,
                user_id=str(callback_query.from_user.id),
                who_invite=ref_code
            )
            await create_user_history(
                user_id=str(callback_query.from_user.id)
            )
        await StartWallet.type_bank.set()
        cached_data[str(callback_query.from_user.id)] = callback_query
        await callback_query.message.edit_text(
            text='Выберите способ оплаты',
            reply_markup=start_skip_wallet(start)
        )
    else:
        await callback_query.message.edit_text(
            text='Подпишитесь на каналы: @asdasdsdfsv',
            reply_markup=check_sub()
        )


@user_dp.callback_query_handler(lambda c:'bank_wallet' in c.data, state=StartWallet.type_bank)
@block_check
async def user_start_wallet_type_bank(callback_query: types.CallbackQuery, state: FSMContext):
    type_bank = callback_query.data.split(':')[1]
    async with state.proxy() as data:
        data["type_bank"] = type_bank
    cached_data[str(callback_query.from_user.id)] = callback_query
    if data['type_bank'] != 'usdt':
        await callback_query.message.edit_text(
            text=f'Введите номер {"счета" if data["type_bank"] == "umoney" else "банковской карты"}:\n\nПример: 5536 6754 5345 5346  (16 цифр)',
            reply_markup=reset_wallet()
        )
        await StartWallet.next()
    else:
        await callback_query.message.edit_text(
            text='Введите адрес своего USDT кошелька в сети BSC [BEP20] (0x...):',
            reply_markup=reset_wallet()
        )
        await StartWallet.next()
    

@user_dp.message_handler(state=StartWallet.card_number)
@block_check
async def user_start_wallet_card_number(msg: types.Message, state: FSMContext):
    callback = cached_data[str(msg.from_user.id)]
    async with state.proxy() as data:
        if data["type_bank"] != 'usdt':
            if len(msg.text) == 19:
                data["card_number"] = msg.text
                await msg.delete()
                match data['type_bank']:
                    case 'rus':
                        await callback.message.edit_text(
                            text='Введите название банка:\n\nПример: Сбербанк',
                            reply_markup=reset_wallet()
                        )
                        await StartWallet.next()
                    case 'ukr':
                        await callback.message.edit_text(
                            text='Введите название банка:\n\nПример: Монобанк',
                            reply_markup=reset_wallet()
                        )
                        await StartWallet.next()
                    case 'umoney':
                        await callback.message.edit_text(
                            text='Введите номер телефона:\n\nПример: +78005553535 (номер к которому привязан банк)',
                            reply_markup=start_wallet_skip_phone()
                        )
                        data["bank_name"] = ''
                        await StartWallet.phone_number.set()

                
            else:
                await msg.delete()
                random_number1 = str(random.randint(1000, 9999))
                random_number2 = str(random.randint(1000, 9999))
                await callback.message.edit_text(
                    text=f'Вы допустили ошибку,\nПример: 5536 {random_number2} {random_number1} 5346  (16 цифр)',
                    reply_markup=reset_wallet()
                )
                return
        else:
            if msg.text[:2] == '0x':
                data["card_number"] = msg.text
                data["bank_name"] = ''
                data["phone_number"] = ''
                await adding_wallet(str(msg.from_user.id), data["type_bank"], data['card_number'], data['bank_name'], data['phone_number'])
                await msg.delete()
                await state.finish()
                await user_main_menu(callback)
            else:
                try:
                    await msg.delete()
                    await callback.message.edit_text(
                        text='⁨⚠️⁩ Ошибка: ⁨некорректный формат кошелька⁩. Попробуйте еще раз.\nВведите адрес своего USDT кошелька в сети Arbitrum One (0x...):',
                        reply_markup=reset_wallet()
                    )
                    return
                except:
                    pass

@user_dp.message_handler(state=StartWallet.bank_name)
@block_check
async def user_start_wallet_bank_name(msg: types.Message, state: FSMContext):
    callback = cached_data[str(msg.from_user.id)]
    async with state.proxy() as data:
        data["bank_name"] = msg.text
    match data["type_bank"]:
        case 'rus':
            await msg.delete()
            await callback.message.edit_text(
                text='Введите Сбп (cистема быстрым платежей)\n\nПример: +78005553535 (номер к которому привязан банк)',
                reply_markup=start_wallet_skip_phone()
            )
            await StartWallet.next()
        case 'ukr':
            async with state.proxy() as data:
                data["phone_number"] = ''
            await msg.delete()
            await adding_wallet(str(msg.from_user.id), data["type_bank"], data["card_number"], data["bank_name"], data["phone_number"])
            await state.finish()
            await user_main_menu(callback)

@user_dp.message_handler(state=StartWallet.phone_number)
@block_check
async def user_start_wallet_phone(msg: types.Message, state: FSMContext):
    callback = cached_data[str(msg.from_user.id)]
    if msg.text[0] == "+":
        async with state.proxy() as data:
            data["phone_number"] = msg.text
        await adding_wallet(str(msg.from_user.id), data["type_bank"], data["card_number"], data["bank_name"], data["phone_number"])
        await state.finish()
        await msg.delete()
        await user_main_menu(callback)
    else:
        await msg.delete()
        await callback.message.edit_text(
            text='Вы допустили ошибку,\n\nПример: +78005553535 (номер к которому привязан банк)',
            reply_markup=start_wallet_skip_phone()
        )
        return

@user_dp.callback_query_handler(lambda c: c.data == 'start_wallet_skip_phone', state=StartWallet.phone_number)
@block_check
async def user_start_wallet_phone_skip(callback_query: types.CallbackQuery, state: FSMContext):
    cached_data[str(callback_query.from_user.id)] = callback_query
    async with state.proxy() as data:
        data["phone_number"] = ''
    await state.finish()
    await adding_wallet(str(callback_query.from_user.id), data["type_bank"], data["card_number"], data["bank_name"], data["phone_number"])
    await user_main_menu(callback_query)

@user_dp.callback_query_handler(lambda c: c.data == 'reser_wallet', state=StartWallet)
@block_check
async def user_start_wallet_reset(callback_query: types.CallbackQuery, state: FSMContext):
    cached_data[str(callback_query.from_user.id)] = callback_query
    await state.finish()
    await user_check_sub_start(callback_query, start=False)

@user_dp.callback_query_handler(lambda c: c.data == 'start_skip_wallet', state=StartWallet.type_bank)
@block_check
async def user_skip_start_wallet(callback_query: types.CallbackQuery, state: FSMContext):
    cached_data[str(callback_query.from_user.id)] = callback_query
    await state.finish()
    await user_main_menu(callback_query, first=True)

#######

####### Мой профиль
@block_check
async def user_main_menu(callback_query: types.CallbackQuery, has_register=False, first=False):
        user_datas = await get_user_data(str(callback_query.from_user.id))
        link = await get_start_link(str(callback_query.from_user.id), encode=True)
        if str(callback_query.from_user.id) in main_menu_icon:
            try:
                await user_bot.delete_message(callback_query.from_user.id, main_menu_icon[str(callback_query.from_user.id)].message_id)
            except:
                pass
        message = await user_bot.send_message(
                callback_query.from_user.id,
                '💵',
                reply_markup=main_kb()
            )
        cached_data[str(callback_query.from_user.id)] = callback_query
        
        main_menu_icon[str(callback_query.from_user.id)] = message
        if has_register:
            await user_bot.send_message(
                callback_query.from_user.id,
                f'Профиль <code>{user_datas["user_id"]}</code>\n{"🛑 У вас не привязан кошелек! Чтобы получать выплаты за выполнение заданий, привяжите его в разделе Кошелек." if user_datas["type_bank"]=="" else "💵 Баланс: " + str(user_datas["balance"])}\n👥 Рефералов: {user_datas["ref_invitees"]}\n\n🔗 Ваша реферальная ссылка:\n<code>{link}</code>',
                parse_mode='HTML',
                reply_markup=inline_main_kb()
            )
        else:
            await callback_query.message.edit_text(
                text=f'Профиль <code>{user_datas["user_id"]}</code>\n{"🛑 У вас не привязан кошелек! Чтобы получать выплаты за выполнение заданий, привяжите его в разделе Кошелек." if user_datas["type_bank"]=="" else "💵 Баланс: " + str(user_datas["balance"])}\n👥 Рефералов: {user_datas["ref_invitees"]}\n\n🔗 Ваша реферальная ссылка:\n<code>{link}</code>',
                parse_mode='HTML',
                reply_markup=inline_main_kb()
            )
            
#######

####### Кошелек

@user_dp.callback_query_handler(lambda c: c.data == 'wallet')
@block_check
async def user_wallet(callback_query: types.CallbackQuery):
    user_datas = await get_user_data(str(callback_query.from_user.id))
    cached_data[str(callback_query.from_user.id)] = callback_query
    if user_datas["type_bank"] == '':
        await user_check_sub_start(callback_query, start=False)
        
    else:
        match user_datas["type_bank"]:
            case 'rus':
                await callback_query.message.edit_text(
                    text=f'Номер банковской карты:\n{user_datas["card_number"]}\n\nБанк: {user_datas["bank_name"]}\n\n СБП: {user_datas["phone_number"]}',
                    reply_markup=edit_wallet()
                )
            case 'ukr':
                await callback_query.message.edit_text(
                    text=f'Номер банковской карты:\n{user_datas["card_number"]}\n\nБанк: {user_datas["bank_name"]}',
                    reply_markup=edit_wallet()
                )
            case 'umoney':
                await callback_query.message.edit_text(
                    text=f'Номер счета:\n{user_datas["card_number"]}\n\nНомер телефона:\n{user_datas["phone_number"]}',
                    reply_markup=edit_wallet()
                )
            case 'usdt':
                await callback_query.message.edit_text(
                    text=f'Привязанный кошелек:\n{user_datas["card_number"]}',
                    reply_markup=edit_wallet()
                )

@user_dp.callback_query_handler(lambda c: c.data == 'wallet_back')
@block_check
async def user_wallet_back(callback_query: types.CallbackQuery):
    cached_data[str(callback_query.from_user.id)] = callback_query
    await user_main_menu(callback_query)

@user_dp.callback_query_handler(lambda c: c.data == 'edit_wallet')
@block_check
async def user_wallet_edit(callback_query: types.CallbackQuery):
    cached_data[str(callback_query.from_user.id)] = callback_query
    await user_check_sub_start(callback_query, start=False)

@user_dp.callback_query_handler(lambda c: c.data == 'delete_wallet')
@block_check
async def user_wallet_delete(callback_query: types.CallbackQuery):
    cached_data[str(callback_query.from_user.id)] = callback_query
    await callback_query.message.edit_text(
        text='Подтвердите удаление кошелька:',
        reply_markup=delete_wallet()
    )

@user_dp.callback_query_handler(lambda c: c.data == 'delete_wallet_back')
@block_check
async def user_wallet_delete_back(callback_query: types.CallbackQuery):
    cached_data[str(callback_query.from_user.id)] = callback_query
    await user_main_menu(callback_query)
    
@user_dp.callback_query_handler(lambda c: c.data == 'delete_wallet_conf')
@block_check
async def user_wallet_delete_conf(callback_query: types.CallbackQuery):
    cached_data[str(callback_query.from_user.id)] = callback_query
    await adding_wallet(str(callback_query.from_user.id), '', '', '', '')
    await user_main_menu(callback_query)

#######

####### Уведомления

@user_dp.callback_query_handler(lambda c: c.data == 'notifications')
@block_check
async def user_notifications(callback_query: types.CallbackQuery):
    cached_data[str(callback_query.from_user.id)] = callback_query
    user_datas = await get_user_data(str(callback_query.from_user.id))
    await callback_query.message.edit_text(
        text='📢 Уведомления\n\nДля изменения параметра нажмите на кнопку снизу.',
        reply_markup=notifications_kb(user_datas["notifications"])
    )

@user_dp.callback_query_handler(lambda c: c.data == 'change_notifications_news')
@block_check
async def user_notifications_news(callback_query: types.CallbackQuery):
    cached_data[str(callback_query.from_user.id)] = callback_query
    user_datas = await get_user_data(str(callback_query.from_user.id))
    notifications = user_datas["notifications"]
    if notifications[0] == '1':
        await change_notifications(str(callback_query.from_user.id), '0'+notifications[1]+notifications[2])
    else:
        await change_notifications(str(callback_query.from_user.id), '1'+notifications[1]+notifications[2])
    await user_notifications(callback_query)

@user_dp.callback_query_handler(lambda c: c.data == 'change_notifications_tasks')
@block_check
async def user_notifications_tasks(callback_query: types.CallbackQuery):
    cached_data[str(callback_query.from_user.id)] = callback_query
    user_datas = await get_user_data(str(callback_query.from_user.id))
    notifications = user_datas["notifications"]
    if notifications[1] == '1':
        await change_notifications(str(callback_query.from_user.id), notifications[0]+'0'+notifications[2])
    else:
        await change_notifications(str(callback_query.from_user.id), notifications[0]+'1'+notifications[2])
    await user_notifications(callback_query)

@user_dp.callback_query_handler(lambda c: c.data == 'change_notifications_assessments')
@block_check
async def user_notifications_assessments(callback_query: types.CallbackQuery):
    cached_data[str(callback_query.from_user.id)] = callback_query
    user_datas = await get_user_data(str(callback_query.from_user.id))
    notifications = user_datas["notifications"]
    if notifications[2] == '1':
        await change_notifications(str(callback_query.from_user.id), notifications[0]+notifications[1]+'0')
    else:
        await change_notifications(str(callback_query.from_user.id), notifications[0]+notifications[1]+'1')
    await user_notifications(callback_query)

@user_dp.callback_query_handler(lambda c: c.data == 'notifications_back')
@block_check
async def user_notifications_back(callback_query: types.CallbackQuery):
    cached_data[str(callback_query.from_user.id)] = callback_query
    await user_main_menu(callback_query)

#######

####### Контакты

@user_dp.callback_query_handler(lambda c: c.data == 'contacts')
@block_check
async def user_contacts(callback_query: types.CallbackQuery):
    cached_data[str(callback_query.from_user.id)] = callback_query
    users_datas = await get_user_data(str(callback_query.from_user.id))
    if users_datas["youtube"] == "":
        await callback_query.message.edit_text(
            text='🔗 Мои аккаунты\n\n⁨Вы еще не добавили ни одного аккаунта.⁩',
            reply_markup=contacnts_no_acc()
        )
    else:
        await callback_query.message.edit_text(
            text=f'🔗 Мои аккаунты\n\nYouTube: {users_datas["youtube"]}',
            reply_markup=contacts_kb(),
            disable_web_page_preview=True
        )

@user_dp.callback_query_handler(lambda c: c.data == 'contacts_back')
@block_check
async def user_contacts_back(callback_query: types.CallbackQuery):
    cached_data[str(callback_query.from_user.id)] = callback_query
    await user_main_menu(callback_query)

#######

####### Мои соцсети

@user_dp.callback_query_handler(lambda c: c.data == 'adding_acc')
@block_check
async def user_contacts_adding(callback_query: types.CallbackQuery):
    cached_data[str(callback_query.from_user.id)] = callback_query
    await callback_query.message.edit_text(
        text='Введите ссылку на аккаунт YouTube',
        reply_markup=adding_youtube()
    )
    await AddingYoutube.link.set()

@user_dp.message_handler(state=AddingYoutube.link)
@block_check
async def user_contacts_adding_youtube_state(msg: types.Message, state: FSMContext):
    callback = cached_data[str(msg.from_user.id)]
    async with state.proxy() as data:
        data["link"] = msg.text
    await change_youtube(str(msg.from_user.id), data['link'])
    await state.finish()
    await msg.delete()
    await user_contacts(callback)

@user_dp.callback_query_handler(lambda c: c.data == 'contacts_youtube_back', state=AddingYoutube.link)
@block_check
async def user_contacts_adding_youtube_back(callback_query: types.CallbackQuery, state: FSMContext):
    cached_data[str(callback_query.from_user.id)] = callback_query
    await state.finish()
    await user_main_menu(callback_query)

#######

####### Помощь


@user_dp.message_handler(Text('🛟 Помощь'), state='*')
@block_check
async def user_support(msg: types.Message, state: FSMContext):
    current_state = await state.get_state()
    if current_state is not None:  # Проверяем, находится ли пользователь в каком-либо состоянии
        await state.finish()  # Завершаем текущее состояние
    callback = cached_data[str(msg.from_user.id)]
    await msg.delete()
    await callback.message.edit_text(
        text='С помощью кнопок ниже вы можете получить ответы на частые вопросы и ознакомиться с полным гайдом по использованию бота, а если у вас останутся вопросы, обратитесь в поддержку.',
        reply_markup=support_kb()
    )


#######

@user_dp.message_handler(Text('👤 Мой профиль'), state='*')
@block_check
async def user_my_profile(msg: types.Message, state: FSMContext):
    current_state = await state.get_state()
    if current_state is not None:  # Проверяем, находится ли пользователь в каком-либо состоянии
        await state.finish()  # Завершаем текущее состояние
    callback = cached_data[str(msg.from_user.id)]
    await msg.delete()
    await user_main_menu(callback)

#######

####### Список заданий

@user_dp.message_handler(Text('❗ Список заданий'), state='*')# Выобр категории в списке заданий
@block_check
async def user_task_list(msg: types.Message, state: FSMContext):
    current_state = await state.get_state()
    if current_state is not None:  # Проверяем, находится ли пользователь в каком-либо состоянии
        await state.finish()  # Завершаем текущее состояние
    callback = cached_data[str(msg.from_user.id)]
    await msg.delete()
    await callback.message.edit_text(
        text='Выберите категорию:',
        reply_markup=await get_all_category_kb()
    )

## Навигация в списке заданий
@user_dp.callback_query_handler(lambda c: 'task_category' in c.data)
@block_check
async def user_all_tasks_in_category(callback_query: types.CallbackQuery):
    category = callback_query.data.split(':')[1]
    tasks_in_category = await get_all_task_in_category(category)
    users_active_tasks = await get_users_active_tasks(str(callback_query.from_user.id), active_tasks=True)
    users_active_tasks = users_active_tasks.split(' ')
    if str(tasks_in_category[0].id) in users_active_tasks:
        if callback_query.from_user.id in tasks_progress[tasks_in_category[0].id]['users']['checking']:
            await callback_query.message.edit_text(
            text=f'🏷️ Категория "{category}"\n\n❗ Задание #{tasks_in_category[0].id}\n\n{tasks_in_category[0].full_text}\n\n💸 Награда: {tasks_in_category[0].price}₽\n\n⏱️Время на выполнение: {tasks_in_category[0].timer + " мин" if tasks_in_category[0].timer != "00" else "бессрочно"}\n\n✏️ Создано: {tasks_in_category[0].start_date}',
            parse_mode='HTML',
            reply_markup=next_task_kb(place=0, category=category, task_number=tasks_in_category[0].id, checking=True)   
        )
        else:
            start_time = tasks_progress[tasks_in_category[0].id]['users']['in_process'][callback_query.from_user.id]
            remaining_time = await is_time_remaining(start_time=start_time, duration_minutes=tasks_in_category[0].timer)
            await callback_query.message.edit_text(
                text=f'🏷️ Категория "{category}"\n\n❗ Задание #{tasks_in_category[0].id}\n\n{tasks_in_category[0].full_text}\n\n💸 Награда: {tasks_in_category[0].price}₽\n\n⏱️Время на выполнение: {tasks_in_category[0].timer + " мин" if tasks_in_category[0].timer != "00" else "бессрочно"}\n\n⌛️ Осталось времени: {str(int(remaining_time)) + " мин" if remaining_time != "Бессрочно" else remaining_time} \n\n✏️ Создано: {tasks_in_category[0].start_date}',
                reply_markup=next_task_kb(place=0, category=category, is_hand=True, task_number=tasks_in_category[0].id)
            )
    else:
        if callback_query.from_user.id in tasks_progress[tasks_in_category[0].id]['users']['done']:
            await callback_query.message.edit_text(
                text=f'🏷️ Категория "{category}"\n\n❗ Задание #{tasks_in_category[0].id}\n\n{tasks_in_category[0].full_text}\n\n💸 Награда: {tasks_in_category[0].price}₽\n\n⏱️Время на выполнение: {tasks_in_category[0].timer + " мин" if tasks_in_category[0].timer != "00" else "бессрочно"}\n\n✏️ Создано: {tasks_in_category[0].start_date}',
                parse_mode='HTML',
                reply_markup=next_task_kb(place=0, category=category, task_number=tasks_in_category[0].id, done=True)   
            )
        elif callback_query.from_user.id in tasks_progress[tasks_in_category[0].id]['users']['rejected']:
            await callback_query.message.edit_text(
                text=f'🏷️ Категория "{category}"\n\n❗ Задание #{tasks_in_category[0].id}\n\n{tasks_in_category[0].full_text}\n\nПричина отколнения: {tasks_progress[tasks_in_category[0].id]["users"]["rejected"][callback_query.from_user.id]["reason"]}\n\n💸 Награда: {tasks_in_category[0].price}₽\n\n⏱️Время на выполнение: {tasks_in_category[0].timer + " мин" if tasks_in_category[0].timer != "00" else "бессрочно"}\n\n✏️ Создано: {tasks_in_category[0].start_date}',
                parse_mode='HTML',
                reply_markup=next_task_kb(place=0, category=category, task_number=tasks_in_category[0].id, rejected=True)   
            )
        else:
            await callback_query.message.edit_text(
                text=f'🏷️ Категория "{category}"\n\n❗ Задание #{tasks_in_category[0].id}\n\n{tasks_in_category[0].full_text}\n\n💸 Награда: {tasks_in_category[0].price}₽\n\n⏱️Время на выполнение: {tasks_in_category[0].timer + " мин" if tasks_in_category[0].timer != "00" else "бессрочно"}\n\n✏️ Создано: {tasks_in_category[0].start_date}',
                parse_mode='HTML',
                reply_markup=next_task_kb(place=0, category=category, task_number=tasks_in_category[0].id)   
            )

@user_dp.callback_query_handler(lambda c: 'next_task' in c.data)
@block_check
async def user_all_tasks_in_category_next(callback_query: types.CallbackQuery):
    place = int(callback_query.data.split(':')[1])
    category = callback_query.data.split(':')[2]
    tasks_in_category = await get_all_task_in_category(category)
    users_active_tasks = await get_users_active_tasks(str(callback_query.from_user.id), active_tasks=True)
    users_active_tasks = users_active_tasks.split(' ')
    if ((len(tasks_in_category) - 1) < place):
        place = 0
    if str(tasks_in_category[place].id) in users_active_tasks:
        if callback_query.from_user.id in tasks_progress[tasks_in_category[place].id]['users']['checking']:
            await callback_query.message.edit_text(
                text=f'🏷️ Категория "{category}"\n\n❗ Задание #{tasks_in_category[place].id}\n\n{tasks_in_category[place].full_text}\n\n💸 Награда: {tasks_in_category[place].price}₽\n\n⏱️ Время на выполнение: {tasks_in_category[place].timer + " мин" if tasks_in_category[place].timer != "00" else "бессрочно"}\n\n✏️ Создано: {tasks_in_category[place].start_date}',
                parse_mode='HTML',
                reply_markup=next_task_kb(place=place, category=category, task_number=tasks_in_category[place].id, checking=True)   
            )
        else:
            start_time = tasks_progress[tasks_in_category[place].id]['users']['in_process'][callback_query.from_user.id]
            remaining_time = await is_time_remaining(start_time=start_time, duration_minutes=tasks_in_category[place].timer)
            await callback_query.message.edit_text(
                text=f'🏷️ Категория "{category}"\n\n❗ Задание #{tasks_in_category[place].id}\n\n{tasks_in_category[place].full_text}\n\n💸 Награда: {tasks_in_category[place].price}₽\n\n⏱️ Время на выполнение: {tasks_in_category[place].timer + " мин" if tasks_in_category[place].timer != "00" else "бессрочно"}\n\n⌛️ Осталось времени: {str(int(remaining_time)) + " мин" if remaining_time != "Бессрочно" else remaining_time}\n\n✏️ Создано: {tasks_in_category[place].start_date}',
                reply_markup=next_task_kb(place=place, category=category, is_hand=True, task_number=tasks_in_category[place].id)
            )
    else:
        if callback_query.from_user.id in tasks_progress[tasks_in_category[place].id]['users']['done']:
            await callback_query.message.edit_text(
                text=f'🏷️ Категория "{category}"\n\n❗ Задание #{tasks_in_category[place].id}\n\n{tasks_in_category[place].full_text}\n\n💸 Награда: {tasks_in_category[place].price}₽\n\n⏱️Время на выполнение: {tasks_in_category[place].timer + " мин" if tasks_in_category[place].timer != "00" else "бессрочно"}\n\n✏️ Создано: {tasks_in_category[place].start_date}',
                parse_mode='HTML',
                reply_markup=next_task_kb(place=place, category=category, task_number=tasks_in_category[place].id, done=True)   
            )
        elif callback_query.from_user.id in tasks_progress[tasks_in_category[place].id]['users']['rejected']:
            await callback_query.message.edit_text(
                text=f'🏷️ Категория "{category}"\n\n❗ Задание #{tasks_in_category[place].id}\n\n{tasks_in_category[place].full_text}\n\nПричина отколнения: {tasks_progress[tasks_in_category[place].id]["users"]["rejected"][callback_query.from_user.id]["reason"]}\n\n💸 Награда: {tasks_in_category[place].price}₽\n\n⏱️Время на выполнение: {tasks_in_category[place].timer + " мин" if tasks_in_category[place].timer != "00" else "бессрочно"}\n\n✏️ Создано: {tasks_in_category[place].start_date}',
                parse_mode='HTML',
                reply_markup=next_task_kb(place=place, category=category, task_number=tasks_in_category[place].id, rejected=True)   
            )
        else:
            await callback_query.message.edit_text(
                text=f'🏷️ Категория "{category}"\n\n❗ Задание #{tasks_in_category[place].id}\n\n{tasks_in_category[place].full_text}\n\n💸 Награда: {tasks_in_category[place].price}₽\n\n⏱️Время на выполнение: {tasks_in_category[place].timer + " мин" if tasks_in_category[place].timer != "00" else "бессрочно"}\n\n✏️ Создано: {tasks_in_category[place].start_date}',
                parse_mode='HTML',
                reply_markup=next_task_kb(place=place, category=category, task_number=tasks_in_category[place].id)   
            )
    
@user_dp.callback_query_handler(lambda c: 'last_task' in c.data)
@block_check
async def user_all_tasks_in_category_last(callback_query: types.CallbackQuery):
    place = int(callback_query.data.split(':')[1])
    category = callback_query.data.split(':')[2]
    tasks_in_category = await get_all_task_in_category(category)
    users_active_tasks = await get_users_active_tasks(str(callback_query.from_user.id), active_tasks=True)
    users_active_tasks = users_active_tasks.split(' ')
    if place == -1:
        place = len(tasks_in_category) - 1
    if str(tasks_in_category[place].id) in users_active_tasks:
        if callback_query.from_user.id in tasks_progress[tasks_in_category[place].id]['users']['checking']:
            await callback_query.message.edit_text(
                text=f'🏷️ Категория "{category}"\n\n❗ Задание #{tasks_in_category[place].id}\n\n{tasks_in_category[place].full_text}\n\n💸 Награда: {tasks_in_category[place].price}₽\n\n⏱️ Время на выполнение: {tasks_in_category[place].timer + " мин" if tasks_in_category[place].timer != "00" else "бессрочно"}\n\n✏️ Создано: {tasks_in_category[place].start_date}',
                parse_mode='HTML',
                reply_markup=next_task_kb(place=place, category=category, task_number=tasks_in_category[place].id, checking=True)   
            )
        else:
            start_time = tasks_progress[tasks_in_category[place].id]['users']['in_process'][callback_query.from_user.id]
            remaining_time = await is_time_remaining(start_time=start_time, duration_minutes=tasks_in_category[place].timer)
            await callback_query.message.edit_text(
                text=f'🏷️ Категория "{category}"\n\n❗ Задание #{tasks_in_category[place].id}\n\n{tasks_in_category[place].full_text}\n\n💸 Награда: {tasks_in_category[place].price}₽\n\n⏱️ Время на выполнение: {tasks_in_category[place].timer + " мин" if tasks_in_category[place].timer != "00" else "бессрочно"}\n\n⌛️ Осталось времени: {str(int(remaining_time)) + " мин" if remaining_time != "Бессрочно" else remaining_time}\n\n✏️ Создано: {tasks_in_category[place].start_date}',
                reply_markup=next_task_kb(place=place, category=category, is_hand=True, task_number=tasks_in_category[place].id)
            )
    else:
        if callback_query.from_user.id in tasks_progress[tasks_in_category[place].id]['users']['done']:
            await callback_query.message.edit_text(
                text=f'🏷️ Категория "{category}"\n\n❗ Задание #{tasks_in_category[place].id}\n\n{tasks_in_category[place].full_text}\n\n💸 Награда: {tasks_in_category[place].price}₽\n\n⏱️Время на выполнение: {tasks_in_category[place].timer + " мин" if tasks_in_category[place].timer != "00" else "бессрочно"}\n\n✏️ Создано: {tasks_in_category[place].start_date}',
                parse_mode='HTML',
                reply_markup=next_task_kb(place=place, category=category, task_number=tasks_in_category[place].id, done=True)   
            )
        elif callback_query.from_user.id in tasks_progress[tasks_in_category[place].id]['users']['rejected']:
            await callback_query.message.edit_text(
                text=f'🏷️ Категория "{category}"\n\n❗ Задание #{tasks_in_category[place].id}\n\n{tasks_in_category[place].full_text}\n\nПричина отколнения: {tasks_progress[tasks_in_category[place].id]["users"]["rejected"][callback_query.from_user.id]["reason"]}\n\n💸 Награда: {tasks_in_category[place].price}₽\n\n⏱️Время на выполнение: {tasks_in_category[place].timer + " мин" if tasks_in_category[place].timer != "00" else "бессрочно"}\n\n✏️ Создано: {tasks_in_category[place].start_date}',
                parse_mode='HTML',
                reply_markup=next_task_kb(place=place, category=category, task_number=tasks_in_category[place].id, rejected=True)   
            )
        else:
            await callback_query.message.edit_text(
                text=f'🏷️ Категория "{category}"\n\n❗ Задание #{tasks_in_category[place].id}\n\n{tasks_in_category[place].full_text}\n\n💸 Награда: {tasks_in_category[place].price}₽\n\n⏱️Время на выполнение: {tasks_in_category[place].timer + " мин" if tasks_in_category[place].timer != "00" else "бессрочно"}\n\n✏️ Создано: {tasks_in_category[place].start_date}',
                parse_mode='HTML',
                reply_markup=next_task_kb(place=place, category=category, task_number=tasks_in_category[place].id)   
            )
##

## приступить к задаче в списке задач
@user_dp.callback_query_handler(lambda c: 'take_task' in c.data)
@block_check
async def user_all_tasks_in_category_take_task(callback_query: types.CallbackQuery):
    task_number = callback_query.data.split(':')[1]
    task = await get_task_datas(int(task_number))
    if task['id'] in tasks_progress:
        if len(tasks_progress[task['id']]['users']['in_process']) + len(tasks_progress[task['id']]['users']['checking']) + len(tasks_progress[task['id']]['users']['done']) + 1 <= int(tasks_progress[task['id']]['limit']):
            await adding_new_activa_task(str(callback_query.from_user.id), task['id'])
            await adding_user_in_process(str(callback_query.from_user.id))
            if callback_query.from_user.id in tasks_progress[task['id']]['users']['canceled']:
                tasks_progress[task['id']]['users']['canceled'].remove(callback_query.from_user.id)
            tasks_progress[task['id']]['users']['in_process'][callback_query.from_user.id] = datetime.datetime.now()
            await user_show_full_task(callback_query, int(task_number))
        else:
            await callback_query.message.edit_text(
                text='Похоже места на выполенение этого задания уже кончились.\n\n Если места появятся, задание появится в "❗ Списке заданий"',
                reply_markup=cancel_task_kb()
            )
    else:
        await callback_query.message.edit_text(
            text='Похоже задание перенесено в архив',
            reply_markup=cancel_task_kb()
        )

##

@user_dp.callback_query_handler(lambda c: c.data == 'back_tasks', state='*')
@block_check
async def user_all_tasks_in_category_back(callback_query: types.CallbackQuery, state: FSMContext):
    current_state = await state.get_state()
    if current_state is not None:  # Проверяем, находится ли пользователь в каком-либо состоянии
        await state.finish()  # Завершаем текущее состояние
    await user_main_menu(callback_query)


####### Кнопки уведомления о новых задачах

@user_dp.callback_query_handler(lambda c: 'reject_new_task' in c.data)
@block_check
async def user_reject_new_task(callback_query: types.CallbackQuery):
    taks_number = callback_query.data.split(':')[1]
    if int(taks_number) in tasks_progress:
        tasks_progress[int(taks_number)]['users']['canceled'].append(callback_query.from_user.id)
    await callback_query.message.delete()    

@block_check
async def user_show_full_task(callback_query: types.CallbackQuery, task_number):
    task = await get_task_datas(int(task_number))
    start_time = tasks_progress[task["id"]]['users']['in_process'][callback_query.from_user.id]
    remaining_time = await is_time_remaining(start_time=start_time, duration_minutes=task["timer"])
    await callback_query.message.edit_text(
        text=f'🏷️ Категория "{task["category"]}"\n\n❗ Задание #{task["id"]}\n\n{task["full_text"]}\n\n💸 Награда: {task["price"]}₽\n\n⏱️ Время на выполнение: {task["timer"] + " мин" if task["timer"] != "00" else "бессрочно"}\n\n⌛️ Осталось времени: {str(int(remaining_time)) + " мин" if remaining_time != "Бессрочно" else remaining_time} \n\n✏️ Создано: {task["start_date"]}',
        reply_markup=new_task_accept_kb(task["id"])
    )
    


@user_dp.callback_query_handler(lambda c: 'accept_new_task' in c.data)
@block_check
async def user_accept_new_task(callback_query: types.CallbackQuery):
    callback = cached_data[str(callback_query.from_user.id)]
    task_number = callback_query.data.split(':')[1]
    task = await get_task_datas(int(task_number))
    if task['id'] in tasks_progress:
        if len(tasks_progress[task['id']]['users']['in_process']) + len(tasks_progress[task['id']]['users']['checking']) + len(tasks_progress[task['id']]['users']['done']) + 1 <= int(tasks_progress[task['id']]['limit']):
            await adding_new_activa_task(str(callback_query.from_user.id), task['id'])
            await adding_user_in_process(str(callback_query.from_user.id))
            await callback_query.message.delete() 
            tasks_progress[task['id']]['users']['in_process'][callback_query.from_user.id] = datetime.datetime.now()
            await user_show_full_task(callback, int(task_number))
        else:
            await callback_query.message.delete()
            await callback.message.edit_text(
                text='Похоже места на выполенение этого задания уже кончились.\n\n Если места появятся, задание появится в "❗ Списке заданий"',
                reply_markup=cancel_task_kb()
            )
    else:
        await callback_query.message.edit_text(
            text='Похоже задание перенесено в архив...'
        )
#######

####### Мои задачи

@user_dp.callback_query_handler(lambda c: c.data == 'my_menu_tasks')# выбор категории в моих задачах
@block_check
async def user_my_tasks(callback_query: types.CallbackQuery):
    my_cat = await get_user_category(str(callback_query.from_user.id))
    await callback_query.message.edit_text(
        text='Выберите категорию:',
        reply_markup=user_my_category_kb(my_cat)
    )

@user_dp.callback_query_handler(lambda c: 'my_tasks' in c.data)# выбор между оцененными и текущими задачами
@block_check
async def user_my_tasks_in_cat(callback_query: types.CallbackQuery):
    cat = callback_query.data.split(':')[1]
    await callback_query.message.edit_text(
        text='Выберите:',
        reply_markup=user_my_tasks_select(cat)
    )


## Навигация в моих оцененных хадачах
@user_dp.callback_query_handler(lambda c: 'my_done_tasks' in c.data)
@block_check
async def user_my_task_history(callback_query: types.CallbackQuery):
    cat = callback_query.data.split(':')[1]
    tasks = await get_users_task_history_by_category(str(callback_query.from_user.id), cat)
    if (len(tasks) != 0) and (tasks != ['']) and (tasks[0] != ''):
        task = await get_task_datas(int(tasks[0]))
        flag1 = task["id"] in tasks_progress
        flag2 = task["id"] in archive_tasks_progerss
        if (flag2 and callback_query.from_user.id in archive_tasks_progerss[task["id"]]["users"]["done"]) or (flag1 and callback_query.from_user.id in tasks_progress[task["id"]]["users"]["done"]):
            await callback_query.message.edit_text(
                text=f'🏷️ Категория "{task["category"]}"\n\n❗ Задание #{task["id"]}\n\n{task["full_text"]}\n\n💸 Награда: {task["price"]}₽\n\n⏱️ Время на выполнение: {task["timer"] + " мин" if task["timer"] != "00" else "бессрочно"}\n\n✏️ Создано: {task["start_date"]}',
                reply_markup=user_task_history_kb(place=0, category=cat, hand=True),
                parse_mode='HTML'
            )
        elif (flag2 and callback_query.from_user.id in archive_tasks_progerss[task["id"]]["users"]["rejected"]) or (flag1 and callback_query.from_user.id in tasks_progress[task["id"]]["users"]["rejected"]):
            await callback_query.message.edit_text(
                text=f'🏷️ Категория "{task["category"]}"\n\n❗ Задание #{task["id"]}\n\n{task["full_text"]}\n\n💸 Награда: {task["price"]}₽\n\n⏱️ Время на выполнение: {task["timer"] + " мин" if task["timer"] != "00" else "бессрочно"}\n\n✏️ Создано: {task["start_date"]}\n\nПричина отколнения: {tasks_progress[task["id"]]["users"]["rejected"][callback_query.from_user.id]["reason"]}',
                reply_markup=user_task_history_kb(place=0, category=cat, hand=False),
                parse_mode='HTML'
            )
    else:
        await callback_query.message.edit_text(
                text='Список пуст'
            )


@user_dp.callback_query_handler(lambda c: 'user_task_history_last' in c.data)
@block_check
async def user_my_task_history_last(callback_query: types.CallbackQuery):
    place = int(callback_query.data.split(':')[1])
    cat = callback_query.data.split(':')[2]
    tasks = await get_users_task_history_by_category(str(callback_query.from_user.id), cat)
    if place == -1:
        place = len(tasks) - 1
    task = await get_task_datas(int(tasks[place]))
    flag1 = task["id"] in tasks_progress
    flag2 = task["id"] in archive_tasks_progerss
    if (flag2 and callback_query.from_user.id in archive_tasks_progerss[task["id"]]["users"]["done"]) or (flag1 and callback_query.from_user.id in tasks_progress[task["id"]]["users"]["done"]):
        await callback_query.message.edit_text(
            text=f'🏷️ Категория "{task["category"]}"\n\n❗ Задание #{task["id"]}\n\n{task["full_text"]}\n\n💸 Награда: {task["price"]}₽\n\n⏱️ Время на выполнение: {task["timer"] + " мин" if task["timer"] != "00" else "бессрочно"}\n\n✏️ Создано: {task["start_date"]}',
            reply_markup=user_task_history_kb(place=place, category=cat, hand=True),
            parse_mode='HTML'
        )
    elif (flag2 and callback_query.from_user.id in archive_tasks_progerss[task["id"]]["users"]["rejected"]) or (flag1 and callback_query.from_user.id in tasks_progress[task["id"]]["users"]["rejected"]):
        await callback_query.message.edit_text(
            text=f'🏷️ Категория "{task["category"]}"\n\n❗ Задание #{task["id"]}\n\n{task["full_text"]}\n\n💸 Награда: {task["price"]}₽\n\n⏱️ Время на выполнение: {task["timer"] + " мин" if task["timer"] != "00" else "бессрочно"}\n\n✏️ Создано: {task["start_date"]}\n\nПричина отколнения: {tasks_progress[task["id"]]["users"]["rejected"][callback_query.from_user.id]["reason"]}',
            reply_markup=user_task_history_kb(place=place, category=cat, hand=False),
            parse_mode='HTML'
        )


@user_dp.callback_query_handler(lambda c: 'user_task_history_next' in c.data)
@block_check
async def user_my_task_history_next(callback_query: types.CallbackQuery):
    place = int(callback_query.data.split(':')[1])
    cat = callback_query.data.split(':')[2]
    tasks = await get_users_task_history_by_category(str(callback_query.from_user.id), cat)
    if ((len(tasks) - 1) < (place)):
        place = 0
    task = await get_task_datas(int(tasks[place]))
    flag1 = task["id"] in tasks_progress
    flag2 = task["id"] in archive_tasks_progerss
    if (flag2 and callback_query.from_user.id in archive_tasks_progerss[task["id"]]["users"]["done"]) or (flag1 and callback_query.from_user.id in tasks_progress[task["id"]]["users"]["done"]):
        await callback_query.message.edit_text(
            text=f'🏷️ Категория "{task["category"]}"\n\n❗ Задание #{task["id"]}\n\n{task["full_text"]}\n\n💸 Награда: {task["price"]}₽\n\n⏱️ Время на выполнение: {task["timer"] + " мин" if task["timer"] != "00" else "бессрочно"}\n\n✏️ Создано: {task["start_date"]}',
            reply_markup=user_task_history_kb(place=place, category=cat, hand=True),
            parse_mode='HTML'
        )
    elif (flag2 and callback_query.from_user.id in archive_tasks_progerss[task["id"]]["users"]["rejected"]) or (flag1 and callback_query.from_user.id in tasks_progress[task["id"]]["users"]["rejected"]):
        await callback_query.message.edit_text(
            text=f'🏷️ Категория "{task["category"]}"\n\n❗ Задание #{task["id"]}\n\n{task["full_text"]}\n\n💸 Награда: {task["price"]}₽\n\n⏱️ Время на выполнение: {task["timer"] + " мин" if task["timer"] != "00" else "бессрочно"}\n\n✏️ Создано: {task["start_date"]}\n\nПричина отколнения: {tasks_progress[task["id"]]["users"]["rejected"][callback_query.from_user.id]["reason"]}',
            reply_markup=user_task_history_kb(place=place, category=cat, hand=False),
            parse_mode='HTML'
        )


##

## Навигация в моих активных задачах
@user_dp.callback_query_handler(lambda c: 'my_active_tasks' in c.data)
@block_check
async def user_my_task_active(callback_query: types.CallbackQuery):
    cat = callback_query.data.split(':')[1]
    tasks  = await get_users_task_active_by_category(str(callback_query.from_user.id), cat, tasks_progress)
    if (len(tasks) != 0) and (tasks != ['']) and (tasks[0] != ''):
        task = await get_task_datas(int(tasks[0]))
        start_time = tasks_progress[task["id"]]['users']['in_process'][callback_query.from_user.id]
        remaining_time = await is_time_remaining(start_time=start_time, duration_minutes=task["timer"])
        await callback_query.message.edit_text(
            text=f'🏷️ Категория "{task["category"]}"\n\n❗ Задание #{task["id"]}\n\n{task["full_text"]}\n\n💸 Награда: {task["price"]}₽\n\n⏱️ Время на выполнение: {task["timer"] + " мин" if task["timer"] != "00" else "бессрочно"}\n\n⌛️ Осталось времени: {str(int(remaining_time)) + " мин" if remaining_time != "Бессрочно" else remaining_time}\n\n✏️ Создано: {task["start_date"]}',
            reply_markup=my_task_active_kb(place=0, category=cat, task_number=task["id"])
        )
    else:
        await callback_query.message.edit_text(
            text='Список пуст'
        )


@user_dp.callback_query_handler(lambda c: 'active_next_my_task' in c.data)
@block_check
async def user_my_task_active_next(callback_query: types.CallbackQuery):
    place = int(callback_query.data.split(':')[1])
    cat = callback_query.data.split(':')[2]
    tasks  = await get_users_task_active_by_category(str(callback_query.from_user.id), cat, tasks_progress)
    if ((len(tasks) - 1) < (place)):
        place = 0
    task = await get_task_datas(int(tasks[place]))
    start_time = tasks_progress[task["id"]]['users']['in_process'][callback_query.from_user.id]
    remaining_time = await is_time_remaining(start_time=start_time, duration_minutes=task["timer"])
    await callback_query.message.edit_text(
        text=f'🏷️ Категория "{task["category"]}"\n\n❗ Задание #{task["id"]}\n\n{task["full_text"]}\n\n💸 Награда: {task["price"]}₽\n\n⏱️ Время на выполнение: {task["timer"] + " мин" if task["timer"] != "00" else "бессрочно"}\n\n⌛️ Осталось времени: {str(int(remaining_time)) + " мин" if remaining_time != "Бессрочно" else remaining_time}\n\n✏️ Создано: {task["start_date"]}',
        reply_markup=my_task_active_kb(place=place, category=cat, task_number=task["id"])
    )

@user_dp.callback_query_handler(lambda c: 'active_last_my_task' in c.data)
@block_check
async def user_my_task_active_last(callback_query: types.CallbackQuery):
    place = int(callback_query.data.split(':')[1])
    cat = callback_query.data.split(':')[2]
    tasks  = await get_users_task_active_by_category(str(callback_query.from_user.id), cat, tasks_progress)
    if place == -1:
        place = len(tasks) - 1
    task = await get_task_datas(int(tasks[place]))
    start_time = tasks_progress[task["id"]]['users']['in_process'][callback_query.from_user.id]
    remaining_time = await is_time_remaining(start_time=start_time, duration_minutes=task["timer"])
    await callback_query.message.edit_text(
        text=f'🏷️ Категория "{task["category"]}"\n\n❗ Задание #{task["id"]}\n\n{task["full_text"]}\n\n💸 Награда: {task["price"]}₽\n\n⏱️ Время на выполнение: {task["timer"] + " мин" if task["timer"] != "00" else "бессрочно"}\n\n⌛️ Осталось времени: {str(int(remaining_time)) + " мин" if remaining_time != "Бессрочно" else remaining_time}\n\n✏️ Создано: {task["start_date"]}',
        reply_markup=my_task_active_kb(place=place, category=cat, task_number=task["id"])
    )

@user_dp.callback_query_handler(lambda c: c.data == 'active_back_tasks')
@block_check
async def user_my_task_active_back(callback_query: types.CallbackQuery):
    cached_data[str(callback_query.from_user.id)] = callback_query
    await user_main_menu(callback_query)
    
##

## Отказ от задачи

@user_dp.callback_query_handler(lambda c: 'cancel_task' in c.data)
@block_check
async def user_cancel_task(callback_query: types.CallbackQuery):
    number_task = callback_query.data.split(':')[1]
    if callback_query.from_user.id in tasks_progress[int(number_task)]['users']['in_process']:
        del tasks_progress[int(number_task)]['users']['in_process'][callback_query.from_user.id]
        tasks_progress[int(number_task)]['users']['rejected'][callback_query.from_user.id] = {
            'reason': 'Отказ'
        }
        await delete_active_task(str(callback_query.from_user.id), number_task)
        await subtract_user_in_process(str(callback_query.from_user.id))
        await adding_user_cancelled(str(callback_query.from_user.id))
        await callback_query.message.edit_text(
            text='Отказ от задачи выполнен!',
            reply_markup=cancel_task_kb()
        )
    else:
        await callback_query.message.edit_text(
            text='Похоже ваше время и так вышло!',
            reply_markup=cancel_task_kb()
        )

##
    

####### 

####### Сдать задачу

@user_dp.callback_query_handler(lambda c: 'hand_task' in c.data)
@block_check
async def user_hand_task(callback_query: types.CallbackQuery, state: FSMContext):
    task_number = callback_query.data.split(':')[1]
    if callback_query.from_user.id not in tasks_progress[int(task_number)]['users']['rejected']:
        await callback_query.message.edit_text(
            text='Отправьте подтверждение (фотографии или видео)',
            reply_markup=confiramtion_file_kb()
        )
        await Confirmation.task_number.set()
        async with state.proxy() as data:
            data["task_number"] = task_number
        await Confirmation.file.set()
    else:
        await callback_query.message.edit_text(
            text='Похоже ваше время и так вышло!',
            reply_markup=cancel_task_kb()
        )



@user_dp.callback_query_handler(lambda c: c.data == 'skip_file_confirmation', state=Confirmation.file)
@block_check
async def user_hand_task_file_state_skip(callback_query: types.CallbackQuery, state: FSMContext):
    await Confirmation.text.set()
    await callback_query.message.edit_text(
        text='Введите сообщение, которое хотите прикрепить к файлам',
        reply_markup=confiramtion_text_kb()
    )


from collections import defaultdict

media_groups = defaultdict(list)

# @block_check
async def download_file(file_id, destination):
    file_info = await user_bot.get_file(file_id)
    file_url = f'https://api.telegram.org/file/bot{user_token}/{file_info.file_path}'
    async with user_bot.session.get(file_url) as response:
        if response.status == 200:
            with open(destination, 'wb') as f:
                f.write(await response.read())



@user_dp.message_handler(content_types=[ContentType.PHOTO, ContentType.VIDEO], state=Confirmation.file)
@block_check
async def user_hand_task_file_state(msg: types.Message, state: FSMContext):
    media_files = []
    message = await user_bot.send_message(
        msg.from_user.id,
        'Файлы загружаются,\n\nНемного подождите...'
    )
    if msg.media_group_id:
        # Добавляем сообщение в медиа-группу
        media_groups[msg.media_group_id].append(msg)
        await asyncio.sleep(1)
        # Ждем все сообщения из медиа-группы
        if len(media_groups[msg.media_group_id]) > 0:
            for media_msg in media_groups[msg.media_group_id]:
                if media_msg.photo:
                    media_files.append(media_msg.photo[-1])
                elif media_msg.video:
                    media_files.append(media_msg.video)
            del media_groups[msg.media_group_id]
    else:
        if msg.photo:
            media_files.append(msg.photo[-1])
        elif msg.video:
            media_files.append(msg.video)

    async with state.proxy() as data:
        task_number = data["task_number"]
        task_folder = f'static/tasks/{task_number}'
        if not os.path.exists(task_folder):
            os.makedirs(task_folder)
        for i, media in enumerate(media_files):
            if isinstance(media, types.PhotoSize):
                file_id = media.file_id
                file_extension = 'jpg'
            else:  # it's a video
                file_id = media.file_id
                file_extension = 'mp4'
            destination = f'{task_folder}/{msg.from_user.id}_{i}.{file_extension}'
            await download_file(file_id=file_id, destination=destination)

    await msg.delete()
    await Confirmation.text.set()
    callback = cached_data[str(msg.from_user.id)]
    await message.delete()
    await callback.message.edit_text(
        text='Введите сообщение, которое хотите прикрепить к файлам',
        reply_markup=confiramtion_text_kb()
    )

@user_dp.callback_query_handler(lambda c: c.data == 'skip_text_confirmation', state=Confirmation.text)
@block_check
async def user_hand_task_text_state_skip(callback_query: types.CallbackQuery, state: FSMContext):
    async with state.proxy() as data:
        data['text'] = ''
    tasks_progress[int(data['task_number'])]['users']['checking'][callback_query.from_user.id] = {
        'text': data['text'],
        'start_date': tasks_progress[int(data['task_number'])]['users']['in_process'][callback_query.from_user.id],
        'end_date': datetime.datetime.now()
    }
    if callback_query.from_user.id in tasks_progress[int(data['task_number'])]['users']['in_process']:
        del tasks_progress[int(data['task_number'])]['users']['in_process'][callback_query.from_user.id]
    await state.finish()
    await callback_query.message.edit_text(
        text='Задание отправлено на проверку',
        reply_markup=cancel_task_kb()
    )


@user_dp.message_handler(state=Confirmation.text)
@block_check
async def user_hand_task_text_state(msg: types.Message, state: FSMContext):
    async with state.proxy() as data:
        data['text'] = msg.text
    tasks_progress[int(data['task_number'])]['users']['checking'][msg.from_user.id] = {
        'text': data['text'],
        'start_date': tasks_progress[int(data['task_number'])]['users']['in_process'][msg.from_user.id],
        'end_date': datetime.datetime.now()
    }
    if msg.from_user.id in tasks_progress[int(data['task_number'])]['users']['in_process']:
        del tasks_progress[int(data['task_number'])]['users']['in_process'][msg.from_user.id]
    await state.finish()
    await msg.delete()
    callback = cached_data[str(msg.from_user.id)]
    await callback.message.edit_text(
        text='Задание отправлено на проверку',
        reply_markup=cancel_task_kb()
    )
    

####### 
####### Поиск задания
@user_dp.message_handler(Text('🔎 Поиск задания'), state='*')
@block_check
async def user_search_task(msg: types.Message, state: FSMContext):
    current_state = await state.get_state()
    if current_state is not None:  # Проверяем, находится ли пользователь в каком-либо состоянии
        await state.finish()  # Завершаем текущее состояние
    callback = cached_data[str(msg.from_user.id)]
    await msg.delete()
    await SearchTask.number_task.set()
    await callback.message.edit_text(
        text='Введите номер задания',
        reply_markup=cancel_task_kb()
    )

@user_dp.message_handler(state=SearchTask.number_task)
@block_check
async def user_search_task_state(msg: types.Message, state: FSMContext):
    callback = cached_data[str(msg.from_user.id)]
    async with state.proxy() as data:
        data['number_task'] = msg.text
    await msg.delete()
    task_data = await get_task_datas(int(data['number_task']))
    await state.finish()
    if task_data:
        if int(data['number_task']) in tasks_progress:
            if msg.from_user.id in tasks_progress[int(data['number_task'])]['users']['in_process']:
                start_time = tasks_progress[task_data["id"]]['users']['in_process'][msg.from_user.id]
                remaining_time = await is_time_remaining(start_time=start_time, duration_minutes=task_data["timer"])
                await callback.message.edit_text(
                    text=f'🏷️ Категория "{task_data["category"]}"\n\n❗ Задание #{task_data["id"]}\n\n{task_data["full_text"]}\n\n💸 Награда: {task_data["price"]}₽\n\n⏱️ Время на выполнение: {task_data["timer"] + " мин" if task_data["timer"] != "00" else "бессрочно"}\n\n⌛️ Осталось времени: {str(int(remaining_time)) + " мин" if remaining_time != "Бессрочно" else remaining_time}\n\n✏️ Создано: {task_data["start_date"]}',
                    reply_markup=search_kb(task_number=int(data['number_task']), in_process=True)
                )
            elif msg.from_user.id in tasks_progress[int(data['number_task'])]['users']['checking']:
                await callback.message.edit_text(
                    text=f'🏷️ Категория "{task_data["category"]}"\n\n❗ Задание #{task_data["id"]}\n\n{task_data["full_text"]}\n\n💸 Награда: {task_data["price"]}₽\n\n⏱️ Время на выполнение: {task_data["timer"] + " мин" if task_data["timer"] != "00" else "бессрочно"}\n\n✏️ Создано: {task_data["start_date"]}',
                    parse_mode='HTML',
                    reply_markup=search_kb(task_number=int(data['number_task']), checking=True)
                )
            elif msg.from_user.id in tasks_progress[int(data['number_task'])]['users']['done']:
                await callback.message.edit_text(
                    text=f'🏷️ Категория "{task_data["category"]}"\n\n❗ Задание #{task_data["id"]}\n\n{task_data["full_text"]}\n\n💸 Награда: {task_data["price"]}₽\n\n⏱️ Время на выполнение: {task_data["timer"] + " мин" if task_data["timer"] != "00" else "бессрочно"}\n\n✏️ Создано: {task_data["start_date"]}',
                    parse_mode='HTML',
                    reply_markup=search_kb(task_number=int(data['number_task']), done=True)
                )
            elif msg.from_user.id in tasks_progress[int(data['number_task'])]['users']['rejected']:
                await callback.message.edit_text(
                    text=f'🏷️ Категория "{task_data["category"]}"\n\n❗ Задание #{task_data["id"]}\n\n{task_data["full_text"]}\n\nПричина отколнения: {tasks_progress[int(data["number_task"])]["users"]["rejected"][msg.from_user.id]["reason"]}\n\n💸 Награда: {task_data["price"]}₽\n\n⏱️ Время на выполнение: {task_data["timer"] + " мин" if task_data["timer"] != "00" else "бессрочно"}\n\n✏️ Создано: {task_data["start_date"]}',
                    parse_mode='HTML',
                    reply_markup=search_kb(task_number=int(data['number_task']), rejected=True)
                )
            else:
                await callback.message.edit_text(
                    text=f'🏷️ Категория "{task_data["category"]}"\n\n❗ Задание #{task_data["id"]}\n\n{task_data["full_text"]}\n\n💸 Награда: {task_data["price"]}₽\n\n⏱️ Время на выполнение: {task_data["timer"] + " мин" if task_data["timer"] != "00" else "бессрочно"}\n\n✏️ Создано: {task_data["start_date"]}',
                    parse_mode='HTML',
                    reply_markup=search_kb(task_number=int(data['number_task']), default=True)
                )
        else:
            if msg.from_user.id in archive_tasks_progerss[int(data['number_task'])]['users']['in_process']:
                start_time = archive_tasks_progerss[task_data["id"]]['users']['in_process'][msg.from_user.id]
                remaining_time = await is_time_remaining(start_time=start_time, duration_minutes=task_data["timer"])
                await callback.message.edit_text(
                    text=f'🏷️ Категория "{task_data["category"]}"\n\n❗ Задание #{task_data["id"]}\n\n{task_data["full_text"]}\n\n💸 Награда: {task_data["price"]}₽\n\n⏱️ Время на выполнение: {task_data["timer"] + " мин" if task_data["timer"] != "00" else "бессрочно"}\n\n⌛️ Осталось времени: {str(int(remaining_time)) + " мин" if remaining_time != "Бессрочно" else remaining_time}\n\n✏️ Создано: {task_data["start_date"]}',
                    reply_markup=search_kb(task_number=int(data['number_task']), in_process=True)
                )
            elif msg.from_user.id in archive_tasks_progerss[int(data['number_task'])]['users']['checking']:
                await callback.message.edit_text(
                    text=f'🏷️ Категория "{task_data["category"]}"\n\n❗ Задание #{task_data["id"]}\n\n{task_data["full_text"]}\n\n💸 Награда: {task_data["price"]}₽\n\n⏱️ Время на выполнение: {task_data["timer"] + " мин" if task_data["timer"] != "00" else "бессрочно"}\n\n✏️ Создано: {task_data["start_date"]}',
                    parse_mode='HTML',
                    reply_markup=search_kb(task_number=int(data['number_task']), checking=True)
                )
            elif msg.from_user.id in archive_tasks_progerss[int(data['number_task'])]['users']['done']:
                await callback.message.edit_text(
                    text=f'🏷️ Категория "{task_data["category"]}"\n\n❗ Задание #{task_data["id"]}\n\n{task_data["full_text"]}\n\n💸 Награда: {task_data["price"]}₽\n\n⏱️ Время на выполнение: {task_data["timer"] + " мин" if task_data["timer"] != "00" else "бессрочно"}\n\n✏️ Создано: {task_data["start_date"]}',
                    parse_mode='HTML',
                    reply_markup=search_kb(task_number=int(data['number_task']), done=True)
                )
            elif msg.from_user.id in archive_tasks_progerss[int(data['number_task'])]['users']['rejected']:
                await callback.message.edit_text(
                    text=f'🏷️ Категория "{task_data["category"]}"\n\n❗ Задание #{task_data["id"]}\n\n{task_data["full_text"]}\n\nПричина отколнения: {archive_tasks_progerss[int(data["number_task"])]["users"]["rejected"][msg.from_user.id]["reason"]}\n\n💸 Награда: {task_data["price"]}₽\n\n⏱️ Время на выполнение: {task_data["timer"] + " мин" if task_data["timer"] != "00" else "бессрочно"}\n\n✏️ Создано: {task_data["start_date"]}',
                    parse_mode='HTML',
                    reply_markup=search_kb(task_number=int(data['number_task']), rejected=True)
                )
            else:
                await callback.message.edit_text(
                    text=f'🏷️ Категория "{task_data["category"]}"\n\n❗ Задание #{task_data["id"]}\n\n{task_data["full_text"]}\n\n💸 Награда: {task_data["price"]}₽\n\n⏱️ Время на выполнение: {task_data["timer"] + " мин" if task_data["timer"] != "00" else "бессрочно"}\n\n✏️ Создано: {task_data["start_date"]}',
                    parse_mode='HTML',
                    reply_markup=search_kb(task_number=int(data['number_task']), default=True)
                )

    else:
        await callback.message.edit_text(
                text='Похоже такого задания не существует...',
                parse_mode='HTML',
                reply_markup=cancel_task_kb()
            )

    

@user_dp.callback_query_handler(lambda c: c.data == 'edit_text_back')
@block_check
async def user_edit_text_back(callback_query: types.CallbackQuery):
    await callback_query.message.delete()



#################################################################################################

############################################### Admin bot ##################################################



from keyboards.admin_keyboards import *
from states.admin_states import *




@user_dp.message_handler(commands='admin', state='*')
@allow_bace_access
async def admin_cmd_start(msg: types.Message, state=None):
    if state:
        current_state = await state.get_state()
        if current_state is not None:  # Проверяем, находится ли пользователь в каком-либо состоянии
            await state.finish()  # Завершаем текущее состояние
    if await has_register(str(msg.from_user.id)):
        await change_is_admin(str(msg.from_user.id), is_admin=True)
    
    await user_bot.send_message(
            msg.from_user.id,
            'Чем я могу помочь?',
            reply_markup=admin_main_kb()
        )
    

####### Добавление новой задачи(админ)

@user_dp.callback_query_handler(lambda c: c.data == 'admin_new_task')
@allow_vip_access
async def admin_new_task(callback_query: types.CallbackQuery):
    await user_bot.send_message(
            callback_query.from_user.id,
            'Введите новую категорию, или выберите из предложенных ниже:',
            reply_markup=await get_category_kb()
        )
    await AdminNewCategory.category.set()

@user_dp.callback_query_handler(lambda c: 'new_task_cat' in c.data, state=AdminNewCategory.category)
@allow_vip_access
async def admin_new_task_callback_category(callback_query: types.CallbackQuery, state: FSMContext):
    category = callback_query.data.split(':')[1]
    async with state.proxy() as data:
        data["category"] = category
    await AdminNewCategory.next()
    await user_bot.send_message(
        callback_query.from_user.id,
        'Введите текст задания',
        reply_markup=new_task_cancel()
    )

@user_dp.message_handler(state=AdminNewCategory.category)
@allow_vip_access
async def admin_new_task_state_category(msg: types.Message, state: FSMContext):
    async with state.proxy() as data:
        data["category"] = msg.text
    await AdminNewCategory.next()
    await user_bot.send_message(
        msg.from_user.id,
        'Введите текст задания',
        reply_markup=new_task_cancel()
    )

@user_dp.message_handler(state=AdminNewCategory.full_text)
@allow_vip_access
async def admin_new_task_state_fulltext(msg: types.Message, state: FSMContext):
    async with state.proxy() as data:
        data["full_text"] = msg.html_text
    await AdminNewCategory.next()
    await user_bot.send_message(
        msg.from_user.id,
        'Введите краткую формулировку',
        reply_markup=new_task_cancel()
    )

@user_dp.message_handler(state=AdminNewCategory.small_text)
@allow_vip_access
async def admin_new_task_state_smalltext(msg: types.Message, state: FSMContext):
    async with state.proxy() as data:
        data["small_text"] = msg.html_text
    await AdminNewCategory.next()
    await user_bot.send_message(
        msg.from_user.id,
        'Введите цену задания',
        reply_markup=new_task_cancel()
    )

@user_dp.message_handler(state=AdminNewCategory.price)
@allow_vip_access
async def admin_new_task_state_price(msg: types.Message, state: FSMContext):
    if msg.text.isdigit():
        async with state.proxy() as data:
            data["price"] = msg.text
        await AdminNewCategory.next()
        await user_bot.send_message(
            msg.from_user.id,
            'Введите время в минутах, если задание бессрочное, введите "00"',
            reply_markup=new_task_cancel()
        )
    else:
        await user_bot.send_message(
            msg.from_user.id,
            'Необходимо ввести число!\n\nВведите цену задания',
            reply_markup=new_task_cancel()
        )
        return

@user_dp.message_handler(state=AdminNewCategory.timer)
@allow_vip_access
async def admin_new_task_state_timer(msg: types.Message, state: FSMContext):
    if msg.text.isdigit():
        async with state.proxy() as data:
            data["timer"] = msg.text
        await AdminNewCategory.next()
        await user_bot.send_message(
            msg.from_user.id,
            'Введите количество людей, которые смогут выполнить задание',
            reply_markup=new_task_cancel()
        )
    else:
        await user_bot.send_message(
            msg.from_user.id,
            'Необходимо ввести число!\n\nВведите время в минутах, если задание бессрочное, введите "00"',
            reply_markup=new_task_cancel()
        )
        return


@user_dp.message_handler(state=AdminNewCategory.count_people)
@allow_vip_access
async def admin_new_task_state_countpeople(msg: types.Message, state: FSMContext):
    if msg.text.isdigit():
        async with state.proxy() as data:
            data["count_people"] = msg.text
        await AdminNewCategory.next()
        await user_bot.send_message(
            msg.from_user.id,
            'так будет выглядеть задание:\n\nВсё верно?'
        )
        await user_bot.send_message(
            msg.from_user.id,
            f'🏷️ Категория: "{data["category"]}"\n\n❗ Задание #{await get_last_task() + 1}\n\nКратко:\n{data["small_text"]}\nПолная формулировка:\n{data["full_text"]}\n\n💸 Награда: {data["price"]}₽\n\n⏱️ Время на выполнение: {data["timer"]  + " мин" if data["timer"] != "00" else "бессрочно"}\n\nЛимит выполнений: {data["count_people"]}',
            reply_markup=new_task_conf(),
            parse_mode='HTML',
            disable_web_page_preview=True
        )
    else:
        await user_bot.send_message(
            msg.from_user.id,
            'Необходимо ввести число!\n\nВведите количество людей, которые смогут выполнить задание',
            reply_markup=new_task_cancel()
        )
        return

@user_dp.callback_query_handler(lambda c: c.data == 'new_task_send', state=AdminNewCategory.confirmation)
@allow_vip_access
async def admin_new_task_state_conf(callback_query: types.CallbackQuery, state: FSMContext):
    async with state.proxy() as data:
        data["confirmation"] = '1'
    list_category = await get_category(list=True)
    if data["category"] not in list_category:
        await adding_category(data["category"])
    await adding_task(
        category=data["category"],
        full_text=data["full_text"],
        small_text=data["small_text"],
        price=data["price"],
        timer=data["timer"],
        count_people=data["count_people"],
        start_date=datetime.datetime.now(),
        who_created=str(callback_query.from_user.id)
    )
    tasks_progress[await get_last_task()] = {
        'start_date': datetime.datetime.now(),
        'timer': data["timer"],
        'limit': data["count_people"],
        'users': {
            'done': {},
            'checking': {},
            'in_process': {},
            'rejected': {},
            'canceled': []
        }
    }
    if not(os.path.exists(f'static/tasks/{await get_last_task()}')):
        os.makedirs(f'static/tasks/{await get_last_task()}')
    if not(os.path.exists(f'static/archive/{await get_last_task()}')):
        os.makedirs(f'static/archive/{await get_last_task()}')
    await adding_stat_count_tasks()
    list_of_users = await get_users_list_for_task()
    async def send_message_to_user(user_id):
        try:
            user_datas = await get_user_data(str(user_id))            
            await user_bot.send_message(
                int(user_id),
                f'Появилось новое задание!\n\nОписание:\n{data["small_text"]}\n\nПриступить к выполнению задания?',
                reply_markup=accept_new_task(await get_last_task()),
                parse_mode='HTML',
                disable_web_page_preview=True
            )
            return True
        except:
            return False

    tasks = [send_message_to_user(user) for user in list_of_users]
    results = await asyncio.gather(*tasks)
    count = sum(results)
    await state.finish()
    await user_bot.send_message(
        callback_query.from_user.id,
        f'Задание отправлено {count} пользователям'
    )
    await admin_cmd_start(callback_query)
    


@user_dp.callback_query_handler(lambda c: c.data == 'new_task_cancel', state=AdminNewCategory)
@allow_vip_access
async def admin_new_task_state_cancel(callback_query: types.CallbackQuery, state: FSMContext):
    await state.finish()
    await admin_cmd_start(callback_query)

####### Все задания

@user_dp.callback_query_handler(lambda c: c.data == 'admin_task')
@allow_bace_access
async def admin_all_tasks(callback_query: types.CallbackQuery):
    await user_bot.send_message(
            callback_query.from_user.id,
            'Выберите',
            reply_markup=admin_all_task_kb()
        )

@user_dp.callback_query_handler(lambda c: c.data == 'active_admin_task')
@allow_bace_access
async def admin_all_task_active(callback_query: types.CallbackQuery):
    await AdminSearchTask.number_task.set()
    await user_bot.send_message(
            callback_query.from_user.id,
            'Выберите или введите номер задания',
            reply_markup=await admin_all_active_tasks_kb(page=0, cached_data=tasks_progress)
        )

@user_dp.callback_query_handler(lambda c: 'admin_all_task_last' in c.data, state=AdminSearchTask.number_task)
@allow_bace_access
async def admin_all_tasks_active_last(callback_query: types.CallbackQuery, state: FSMContext):
    page = int(callback_query.data.split(':')[1])
    tasks = await get_all_tasks()
    if page == -1:
        page = (len(tasks) // 7 + 1) - 1
    await callback_query.message.edit_text(
        text='Выберите или введите номер задания',
        reply_markup=await admin_all_active_tasks_kb(page=page, cached_data=tasks_progress)
    )
    return

@user_dp.callback_query_handler(lambda c: 'admin_all_task_next' in c.data, state=AdminSearchTask.number_task)
@allow_bace_access
async def admin_all_tasks_active_next(callback_query: types.CallbackQuery, state: FSMContext):
    page = int(callback_query.data.split(':')[1])
    tasks = await get_all_tasks()
    if ((len(tasks) // 7 + 1) <= (page)):
        page = 0
    await callback_query.message.edit_text(
        text='Выберите или введите номер задания',
        reply_markup=await admin_all_active_tasks_kb(page=page, cached_data=tasks_progress)
    )
    return

@user_dp.callback_query_handler(lambda c: 'admin_all_tasks' in c.data, state=AdminSearchTask.number_task)
@allow_bace_access
async def admin_all_tasks_active_full(callback_query: types.CallbackQuery, state: FSMContext):
    async with state.proxy() as data:
        data["task_number"] = callback_query.data.split(':')[1]
    await state.finish()
    await admin_show_full_active_task(callback_query, data['task_number'])

@user_dp.message_handler(state=AdminSearchTask.number_task)
@allow_bace_access
async def admin_all_tasks_active_state(msg: types.Message, state: FSMContext):
    async with state.proxy() as data:
        data["task_number"] = msg.text
    await state.finish()
    await admin_show_full_active_task(msg, data['task_number'])

@allow_bace_access
async def admin_show_full_active_task(msg: types.Message, number_task):
    task = await get_task_datas(int(number_task))
    if task:
        await user_bot.send_message(
            msg.from_user.id,
            text=f'🏷️ Категория "{task["category"]}"\n\n❗ Задание #{task["id"]}\n\n{task["full_text"]}\n\n💸 Награда: {task["price"]}₽\n\n⏱️ Время на выполнение: {task["timer"] + " мин" if task["timer"] != "00" else "бессрочно"}\n\n✏️ Создано: {task["start_date"]}',
            parse_mode='HTML',
            reply_markup=admin_show_full_task_kb(number_task=number_task)
        )
    else:
        await user_bot.send_message(
            msg.from_user.id,
            'Похоже такого задания не существует...'
        )

## Изменение задания

@user_dp.callback_query_handler(lambda c: 'show_full_task_edit' in c.data)
@allow_bace_access
async def admin_show_full_active_task_edit(callback_query: types.CallbackQuery, state: FSMContext):
    number_task = callback_query.data.split(':')[1]
    await AdminEditText.number_task.set()
    async with state.proxy() as data:
        data["number_task"] = number_task
    await AdminEditText.text.set()
    await user_bot.send_message(
        callback_query.from_user.id,
        'Отправьте измененный текст задания'
    )

@user_dp.message_handler(state=AdminEditText.text)
@allow_bace_access
async def admin_show_full_active_task_edit_state(msg: types.Message, state: FSMContext):
    async with state.proxy() as data:
        data["text"] = msg.text
    task = await get_task_datas(int(data["number_task"]))
    await user_bot.send_message(
        msg.from_user.id,
        'Задание будет выглядеть вот так:\n\nВсё верно?'
    )
    await user_bot.send_message(
        msg.from_user.id,
        f'🏷️ Категория "{task["category"]}"\n\n❗ Задание #{task["id"]}\n\n{data["text"]}\n\n💸 Награда: {task["price"]}₽\n\n⏱️ Время на выполнение: {task["timer"] + " мин" if task["timer"] != "00" else "бессрочно"}\n\n✏️ Создано: {task["start_date"]}',
        parse_mode='HTML',
        reply_markup=admin_edit_task_text_conf_kb(),
        disable_web_page_preview=True
    )
    await AdminEditText.confirmation.set()

@user_dp.callback_query_handler(lambda c: c.data == 'admin_edit_text_conf', state=AdminEditText.confirmation)
@allow_bace_access
async def admin_show_full_active_task_edit_state_conf(callback_query: types.CallbackQuery, state: FSMContext):
    async with state.proxy() as data:
        await edit_task_text(int(data["number_task"]), data["text"])
    await state.finish()
    message = await user_bot.send_message(
        callback_query.from_user.id,
        'Оповещаем пользователей, немного подождите...'
    )
    list_of_users = []
    for users in tasks_progress[int(data['number_task'])]['users']['in_process']:
        list_of_users.append(users)

    async def send_message_to_user(user_id):
        try:
            await user_bot.send_message(
                int(user_id),
                f'Задание #{data["number_task"]} было изменено,\n\nНовые условия:\n\n{data["text"]}',
                parse_mode='HTML',
                reply_markup=user_edit_text_kb(),
            )
            return True
        except:
            return False

    tasks = [send_message_to_user(user) for user in list_of_users]
    results = await asyncio.gather(*tasks)
    await message.delete()
    await user_bot.send_message(
        callback_query.from_user.id,
        'Задание изменено!'
    )
    await admin_show_full_active_task(callback_query, int(data['number_task']))

##

## Удаление задания

@user_dp.callback_query_handler(lambda c: 'show_full_task_delete' in c.data)
@allow_bace_access
async def admin_delete_task(callback_query: types.CallbackQuery):
    number_task = callback_query.data.split(':')[1]
    await user_bot.send_message(
        callback_query.from_user.id,
        'Вы уверены что хотите удалить задание?',
        reply_markup=admin_delete_task_conf_kb(number_task=number_task)
    )

@user_dp.callback_query_handler(lambda c: 'admin_delete_conf' in c.data)
@allow_bace_access
async def admin_delete_task_conf(callback_query: types.CallbackQuery):
    number_task = callback_query.data.split(':')[1]
    message = await user_bot.send_message(
        callback_query.from_user.id,
        'Данные удаляются, немного подождите...'
    )
    users = [] 
    for user in tasks_progress[int(number_task)]['users']['in_process']:
        users.append(user)
    for user in tasks_progress[int(number_task)]['users']['checking']:
        users.append(user)
    for user in users:
        await delete_active_task(user, int(number_task))
        await delete_history_task(user, int(number_task))
    rejected = 0
    cancelled = len(tasks_progress[int(number_task)]['users']['canceled'])
    times = 0
    done = len(tasks_progress[int(number_task)]['users']['done'])
    for user in tasks_progress[int(number_task)]['users']['rejected']:
        if tasks_progress[int(number_task)]['users']['rejected'][user]['reason'] != 'Лимит времени':
            rejected += 1
        else:
            times += 1
    task_datas = await get_task_datas(int(number_task))
    await adding_archive_task(
        number_task=task_datas["id"],
        category=task_datas["category"],
        full_text=task_datas["full_text"],
        small_text=task_datas["small_text"],
        price=task_datas["price"],
        timer=task_datas["timer"],
        count_people=task_datas["count_people"],
        start_date=task_datas["start_date"],
        end_date=datetime.datetime.now(),
        who_created=task_datas["who_created"],
        rejected=rejected,
        cancelled=cancelled,
        times=times,
        done=done
    )
    await delete_task(int(number_task))
    await adding_stat_del_task()
    async def send_message_to_user(user_id):
        try:
            await user_bot.send_message(
                int(user_id),
                f'Задание #{number_task} было удалено!',
                parse_mode='HTML',
                reply_markup=user_edit_text_kb(),
            )
            return True
        except:
            return False

    tasks = [send_message_to_user(user) for user in users]
    results = await asyncio.gather(*tasks)
    archive_tasks_progerss[int(number_task)] = tasks_progress[int(number_task)]
    del tasks_progress[int(number_task)]
    shutil.rmtree(f'static/tasks/{number_task}')
    await message.delete()
    await user_bot.send_message(
        callback_query.from_user.id,
        'Задание успешно удалено!'
    )
    await admin_all_task_active(callback_query)

##

## Выполнения

admin_checking_cach = {}

@user_dp.callback_query_handler(lambda c: 'show_full_task_checking' in c.data)
@allow_bace_access
async def admin_checking_task(callback_query: types.CallbackQuery, number_task=0, place=0):
    try:
        number_task = callback_query.data.split(':')[1]
    except:
        number_task = number_task
    if len(tasks_progress[int(number_task)]['users']['checking']) != 0:
        user_id = list(tasks_progress[int(number_task)]['users']['checking'].keys())[place]
        user_text = tasks_progress[int(number_task)]['users']['checking'][user_id]['text']
        media = []
        if callback_query.from_user.id in admin_checking_cach:
            del admin_checking_cach[callback_query.from_user.id]
        if os.path.exists(f'static/tasks/{number_task}'):
            all_items = os.listdir(f'static/tasks/{number_task}')
            files = [item for item in all_items if os.path.isfile(os.path.join(f'static/tasks/{number_task}', item))]
            for file in files:
                if str(user_id) in file:
                    if 'mp4' in file:
                        media.append(types.InputMediaVideo(media=open(f'static/tasks/{number_task}/{file}', 'rb'), caption=user_text))
                    elif 'jpg' in file:
                        media.append(types.InputMediaPhoto(media=open(f'static/tasks/{number_task}/{file}', 'rb'), caption=user_text))
            user_data = await get_user_data(str(user_id))

            if media:
                message = await user_bot.send_media_group(
                    callback_query.from_user.id, 
                    media=media,
                )
                await user_bot.send_message(
                    callback_query.from_user.id,
                    f'Username: @{user_data["user_name"]}\n\nId: {user_data["user_id"]}',
                    reply_markup=admin_checking_kb(number_task=number_task, user_id=user_id, place=place)
                )
                admin_checking_cach[callback_query.from_user.id] = message
            else:
                await user_bot.send_message(
                    callback_query.from_user.id,
                    f'Username: @{user_data["user_name"]}\n\nId: {user_data["user_id"]}\n\nПользователь не прикрепил медиафайлов\nКомментарий: {user_text}',
                    reply_markup=admin_checking_kb(number_task=number_task, user_id=user_id, place=place)
                )
    else:
        await user_bot.send_message(
            callback_query.from_user.id,
            'На этом всё!'
        )

# Навигация в проверке
@user_dp.callback_query_handler(lambda c: 'admin_last_checking' in c.data)
@allow_bace_access
async def admin_checking_task_last(callback_query: types.CallbackQuery):
    number_task = callback_query.data.split(':')[1]
    place = int(callback_query.data.split(':')[2])
    users = tasks_progress[int(number_task)]['users']['checking']
    if place == -1:
        place = len(users) - 1
    await callback_query.message.delete()
    if callback_query.from_user.id in admin_checking_cach:
        for message in admin_checking_cach[callback_query.from_user.id]:
            await message.delete()
        del admin_checking_cach[callback_query.from_user.id]
    await admin_checking_task(callback_query, number_task=number_task, place=place)

@user_dp.callback_query_handler(lambda c: 'admin_next_checking' in c.data)
@allow_bace_access
async def admin_checking_task_last(callback_query: types.CallbackQuery):
    number_task = callback_query.data.split(':')[1]
    place = int(callback_query.data.split(':')[2])
    users = tasks_progress[int(number_task)]['users']['checking']
    if ((len(users) - 1) < (place)):
        place = 0
    await callback_query.message.delete()
    if callback_query.from_user.id in admin_checking_cach:
        for message in admin_checking_cach[callback_query.from_user.id]:
            await message.delete()
        del admin_checking_cach[callback_query.from_user.id]
    await admin_checking_task(callback_query, number_task=number_task, place=place)




@user_dp.callback_query_handler(lambda c: 'accept_admin__task' in c.data)
@allow_bace_access
async def admin_checking_task_accept(callback_query: types.CallbackQuery):
    number_task = callback_query.data.split(':')[1]
    user_id = callback_query.data.split(':')[2]
    place = callback_query.data.split(':')[3]
    user_datas = await get_user_data(user_id)
    tasks_progress[int(number_task)]['users']['done'][int(user_id)] = {
        'start_date': tasks_progress[int(number_task)]['users']['checking'][int(user_id)]['start_date'],
        'end_date': tasks_progress[int(number_task)]['users']['checking'][int(user_id)]['end_date'],
        'admin_id': callback_query.from_user.id,
        'admin_name': callback_query.from_user.username 
    }
    if int(user_id) in tasks_progress[int(number_task)]['users']['checking']:
        del tasks_progress[int(number_task)]['users']['checking'][int(user_id)]
    await delete_active_task(user_id, number_task)
    await subtract_user_in_process(user_id)
    await adding_user_done(user_id)
    await adding_new_history_task(user_id, number_task)
    await adding_new_done_task(user_id, number_task)
    if os.path.exists(f'static/tasks/{number_task}'):
        all_items = os.listdir(f'static/tasks/{number_task}')
        files = [item for item in all_items if os.path.isfile(os.path.join(f'static/tasks/{number_task}', item))]
        for file in files:
            if str(user_id) in file:
                shutil.move(f'static/tasks/{number_task}/{file}', f'static/archive/{number_task}/{file}')
    task_datas = await get_task_datas(int(number_task))
    await adding_balance(user_id=user_id, reward=task_datas["price"])
    if user_datas["who_invite"] != '':
        await adding_ref_balance(user_datas["who_invite"], int(int(task_datas["price"]) * 0.15))
    await adding_stat_accepted()
    await callback_query.message.delete()
    if callback_query.from_user.id in admin_checking_cach:
        for message in admin_checking_cach[callback_query.from_user.id]:
            await message.delete()
        del admin_checking_cach[callback_query.from_user.id]
    await admin_checking_task(callback_query, number_task=number_task, place=int(place)-1)



@user_dp.callback_query_handler(lambda c: 'reject_admin_task' in c.data)
@allow_bace_access
async def admin_checking_task_reject(callback_query: types.CallbackQuery, state: FSMContext):
    number_task = callback_query.data.split(':')[1]
    user_id = callback_query.data.split(':')[2]
    place = callback_query.data.split(':')[3]
    await AdminRejectTask.place.set()
    async with state.proxy() as data:
        data['place'] = place
        data["user_id"] = user_id
        data["number_task"] = number_task
    await AdminRejectTask.couse.set()
    await callback_query.message.delete()
    await user_bot.send_message(
        callback_query.from_user.id,
        'Введите причину отклонения',
        reply_markup=admin_reject_task_kb()
    )

@user_dp.callback_query_handler(lambda c: c.data == 'admin_reject_task_cancel', state=AdminRejectTask.couse)
@allow_bace_access
async def admin_checking_task_reject_state_skip(callback_query: types.CallbackQuery, state: FSMContext):
    if callback_query.from_user.id in admin_checking_cach:
        for message in admin_checking_cach[callback_query.from_user.id]:
            await message.delete()
        del admin_checking_cach[callback_query.from_user.id]
    async with state.proxy() as data:
        pass
    await state.finish()
    await admin_checking_task(callback_query, number_task=data['number_task'], place=int(data['place']))

@user_dp.message_handler(state=AdminRejectTask.couse)
@allow_bace_access
async def admin_checking_task_reject_state(msg: types.Message, state: FSMContext):
    async with state.proxy() as data:
        data["couse"] = msg.text
    await state.finish()
    if msg.from_user.id in admin_checking_cach:
        for message in admin_checking_cach[msg.from_user.id]:
            await message.delete()
        del admin_checking_cach[msg.from_user.id]
    tasks_progress[int(data['number_task'])]['users']['rejected'][int(data["user_id"])] = {
        'reason': data['couse'],
        'start_date': tasks_progress[int(data["number_task"])]['users']['checking'][int(data['user_id'])]['start_date'],
        'end_date': tasks_progress[int(data["number_task"])]['users']['checking'][int(data['user_id'])]['end_date'],
        'admin_id': msg.from_user.id,
        'admin_name': msg.from_user.username 
    }
    if os.path.exists(f'static/tasks/{data["number_task"]}'):
        all_items = os.listdir(f'static/tasks/{data["number_task"]}')
        files = [item for item in all_items if os.path.isfile(os.path.join(f'static/tasks/{data["number_task"]}', item))]
        for file in files:
            if str(data["number_task"]) in file:
                shutil.move(f'static/tasks/{data["number_task"]}/{file}', f'static/archive/{data["number_task"]}/{file}')
    if int(data["user_id"]) in tasks_progress[int(data["number_task"])]['users']['checking']:
        del tasks_progress[int(data["number_task"])]['users']['checking'][int(data['user_id'])]
    await delete_active_task(user_id=data['user_id'], task_number=data['number_task'])
    await subtract_user_in_process(data['user_id'])
    await adding_user_rejected(data['user_id'])
    await adding_new_history_task(data['user_id'], data['number_task'])
    await adding_stat_rejected()
    user_datas = await get_user_data(data["user_id"])
    if user_datas["notifications"][2] != '0':
        await user_bot.send_message(
            int(data['user_id']),
            f'Задание #{data["number_task"]} было отклонено,\n\nПричина: {data["couse"]}',
            reply_markup=user_edit_text_kb()
        )
    await state.finish()
    await admin_checking_task(msg, number_task=data['number_task'], place=int(data['place'])-1)

##

#######

####### Архив

@user_dp.callback_query_handler(lambda c: c.data == 'admin_archeve_task')
@allow_bace_access
async def admin_all_task_archive(callback_query: types.CallbackQuery):
    await AdminSearchArchiveTask.number_task.set()
    await user_bot.send_message(
        callback_query.from_user.id,
        'Выберите или введите номер задания',
        reply_markup=await admin_all_archive_tasks_kb(page=0)
    )    

## Навигация в архиве

@user_dp.callback_query_handler(lambda c: 'admin_all_archive_tasks_last' in c.data, state=AdminSearchArchiveTask.number_task)
@allow_bace_access
async def admin_all_task_archive_last(callback_query: types.CallbackQuery):
    page = int(callback_query.data.split(':')[1])
    tasks = await get_all_archive_tasks()
    if page == -1:
        page = (len(tasks) // 7 + 1) - 1
    await callback_query.message.edit_text(
        text='Выберите или введите номер задания',
        reply_markup=await admin_all_archive_tasks_kb(page=page)
    )    
    return


@user_dp.callback_query_handler(lambda c: 'admin_all_archive_tasks_next' in c.data, state=AdminSearchArchiveTask.number_task)
@allow_bace_access
async def admin_all_task_archive_next(callback_query: types.CallbackQuery):
    page = int(callback_query.data.split(':')[1])
    tasks = await get_all_archive_tasks()
    if ((len(tasks) // 7 + 1) <= (page)):
        page = 0
    await callback_query.message.edit_text(
        text='Выберите или введите номер задания',
        reply_markup=await admin_all_archive_tasks_kb(page=page)
    )    
    return

@user_dp.message_handler(state=AdminSearchArchiveTask.number_task)
@allow_bace_access
async def admin_all_task_archive_state(msg: types.Message, state: FSMContext):
    async with state.proxy() as data:
        data["number_task"] = msg.text
    await state.finish()
    await admin_all_task_archive_full(msg, data['number_task'])

@user_dp.callback_query_handler(lambda c: 'admin_all_archive_tasks' in c.data, state=AdminSearchArchiveTask.number_task)
@allow_bace_access
async def admin_all_task_archive_button(callback_query: types.CallbackQuery, state: FSMContext):
    number_task = callback_query.data.split(':')[1]
    await state.finish()
    await admin_all_task_archive_full(callback_query, number_task)

##
@allow_bace_access
async def admin_all_task_archive_full(msg: types.Message, number_task):
    task = await get_archive_task_datas(number_task=int(number_task))
    if task:
        await user_bot.send_message(
            msg.from_user.id,
            text=f'🏷️ Категория "{task["category"]}"\n\n❗ Задание #{task["number_task"]}\n\n{task["full_text"]}\n\n💸 Награда: {task["price"]}₽\n\n⏱️ Время на выполнение: {task["timer"] + " мин" if task["timer"] != "00" else "бессрочно"}\n\n✏️ Создано: {task["start_date"]}',
            parse_mode='HTML'
        )
        await user_bot.send_message(
            msg.from_user.id,
            text=f'Дата создания: {task["start_date"]}\nДата окончания: {task["end_date"]}\nОтклонено куратором: {task["rejected"]}\nОтменено: {task["cancelled"]}\nНе уложились по времени: {task["times"]}',
            reply_markup=admin_all_task_archive_full_kb(number_task)
        )
    

@user_dp.callback_query_handler(lambda c: c.data == 'all_task_archive_full_back')
@allow_bace_access
async def get_archive_task_datas_back(callback_query: types.CallbackQuery):
    await admin_all_task_archive(callback_query)


## Список выполнивших

@user_dp.callback_query_handler(lambda c: 'list_of_done' in c.data)
@allow_bace_access
async def get_archive_task_users(callback_query: types.CallbackQuery):
    number_task = callback_query.data.split(':')[1]
    users = await get_done_task_users(number_task)
    if users:
        workbook = openpyxl.Workbook()
        
        # Выбираем активный лист
        sheet = workbook.active
        
        # Заголовки столбцов
        sheet['A1'] = 'ID'
        sheet['B1'] = 'Name'
        
        # Заполняем столбцы ID и Name
        for index, (user_id, name) in enumerate(users, start=2):
            sheet.cell(row=index, column=1, value=user_id)
            sheet.cell(row=index, column=2, value=name)
        
        # Сохраняем файл
        workbook.save('static/excel/id_names.xlsx')
        await user_bot.send_document(
            callback_query.from_user.id,
            document=open('static/excel/id_names.xlsx', 'rb')
        )
    else:
        await user_bot.send_message(
            callback_query.from_user.id,
            'Похоже никто не выполнил это задание'
        )

##

#######
    
    
####### Пользователи

## Показать информацию

@user_dp.callback_query_handler(lambda c: c.data == 'admin_users')
@allow_bace_access
async def admin_users(callback_query: types.CallbackQuery):
    await AdminSearchUser.user_id.set()
    await user_bot.send_message(
        callback_query.from_user.id,
        'Введите username или id пользователя'
    )

@user_dp.message_handler(state=AdminSearchUser.user_id)
@allow_bace_access
async def admin_users_state(msg: types.Message, state: FSMContext):
    async with state.proxy() as data:
        data["user_id"] = msg.text
    if data['user_id'].isdigit():
        user_datas = await get_user_data(int(data['user_id']))
    else:
        user_datas = await get_user_data_by_username(data['user_id'])
    if user_datas:
        if not(data['user_id'].isdigit()):
            async with state.proxy() as data:
                data["user_id"] = user_datas["user_id"]
        await state.finish()
        warning = ''
        if int(data['user_id']) in warnings:
            warning = ''
            for index, war in enumerate(warnings[int(data['user_id'])], start=1):
                warning += f'{index}: {war}\n'

        await user_bot.send_message(
            msg.from_user.id,
            f'Username: {user_datas["user_name"]}\nId: {user_datas["user_id"]}\nБаланс: {user_datas["balance"]}\nВыполняет заданий: {user_datas["in_process"]}\nВыполнил заданий: {user_datas["done"]}\nОтменил выполнений: {user_datas["cancelled"]}\nОтклонений: {user_datas["rejected"]}\nПринял заданий: {int(user_datas["in_process"]) + int(user_datas["done"]) + int(user_datas["cancelled"]) + int(user_datas["rejected"])}\nОпозданий: {user_datas["times"]}\n{"Предупреждения:  " + warning if len(warning) != 0 else None}',
            reply_markup=admin_users_kb(data['user_id'])
        )


@user_dp.callback_query_handler(lambda c: c.data == 'admin_back_user')
async def admin_users_back(callback_query: types.CallbackQuery):
    await admin_users(callback_query)
    
##

## Предупреждения

@user_dp.callback_query_handler(lambda c: 'admin_user_warning' in c.data)
@allow_bace_access
async def admin_users_warnings(callback_query: types.CallbackQuery):
    user_id = callback_query.data.split(":")[1]
    warning = ''
    if int(user_id) in warnings:
        warning = ''
        for index, war in enumerate(warnings[int(user_id)], start=1):
            warning += f'{index}: {war}\n'
        warn_len = len(warnings[int(user_id)])
    if len(warning) != 0:
        await user_bot.send_message(
            callback_query.from_user.id,
            f'У пользователя уже {warn_len} предупреждений.\n\n{warning}',
            reply_markup=admin_users_warnings_kb(user_id)
        )
    else:
        await user_bot.send_message(
            callback_query.from_user.id,
            'У пользователя 0 предупреждений.',
            reply_markup=admin_users_warnings_kb(user_id)
        )

@user_dp.callback_query_handler(lambda c: 'admin_adding_warning' in c.data)
@allow_bace_access
async def admin_adding_warning(callback_query: types.CallbackQuery, state: FSMContext):
    user_id = callback_query.data.split(':')[1]
    await AdminAddingWarning.user_id.set()
    async with state.proxy() as data:
        data["user_id"] = user_id
    await AdminAddingWarning.text.set()
    await user_bot.send_message(
        callback_query.from_user.id,
        'Введите комментарий к предупреждению:'
    )

@user_dp.message_handler(state=AdminAddingWarning.text)
@allow_bace_access
async def admin_adding_warning_state(msg: types.Message, state: FSMContext):
    async with state.proxy() as data:
        data["text"] = msg.text
    await state.finish()
    if int(data["user_id"]) in warnings:
        warnings[int(data['user_id'])].append(str(datetime.datetime.now()) + ': ' + data['text'])
    else:
        warnings[int(data['user_id'])] = [str(datetime.datetime.now()) + ': ' + data['text']]
    await user_bot.send_message(
        msg.from_user.id,
        'Предупреждение добавлено!'
    )

@user_dp.callback_query_handler(lambda c: 'admin_block_user' in c.data)
@allow_bace_access
async def admin_block_user(callback_query: types.CallbackQuery):
    user_id = callback_query.data.split(':')[1]
    await user_bot.send_message(
        callback_query.from_user.id,
        'Вы уверены что хотите заблокировать пользователя?',
        reply_markup=admin_ublock_user_conf(user_id)
    )


@user_dp.callback_query_handler(lambda c: 'admin_conf_block_user' in c.data)
@allow_bace_access
async def admin_block_user_conf(callback_query: types.CallbackQuery):
    global block_users
    user_id = callback_query.data.split(':')[1]
    await block_user(int(user_id))
    block_users.add(int(user_id))
    await user_bot.send_message(
        callback_query.from_user.id,
        'Пользователь заблокирован'
    )

@user_dp.callback_query_handler(lambda c: c.data == 'admin_conf_calcel_block')
@allow_bace_access
async def admin_block_user_cancel(callback_query: types.CallbackQuery):
    await user_bot.send_message(
        callback_query.from_user.id,
        'Действие отменено',
    )
    await admin_users(callback_query)

##

## Выдать баланс

@user_dp.callback_query_handler(lambda c: 'admin_user_adding_balance' in c.data)
@allow_vip_access
async def admin_adding_user_balance(callback_query: types.CallbackQuery, state: FSMContext):
    user_id = callback_query.data.split(':')[1]
    await AdminAddingBalance.user_id.set()
    async with state.proxy() as data:
        data["user_id"] = user_id
    await AdminAddingBalance.balance.set()
    await user_bot.send_message(
        callback_query.from_user.id,
        'Введите сумму, которую хотите зачислить пользователю:'
    )


@user_dp.message_handler(state=AdminAddingBalance.balance)
@allow_vip_access
async def admin_adding_user_balance_state(msg: types.Message, state: FSMContext):
    async with state.proxy() as data:
        data["balance"] = msg.text
    await state.finish()
    await adding_balance(int(data['user_id']), int(data['balance']))
    await user_bot.send_message(
        msg.from_user.id,
        'Зачисление на баланс выполнено успешно!'
    )
##

## Вычесть баланс

@user_dp.callback_query_handler(lambda c: 'admin_user_subtract_balance' in c.data)
@allow_bace_access
async def admin_subtract_user_balance(callback_query: types.CallbackQuery, state: FSMContext):
    user_id = callback_query.data.split(':')[1]
    await AdminSubtractBalance.user_id.set()
    async with state.proxy() as data:
        data["user_id"] = user_id
    await AdminSubtractBalance.balance.set()
    await user_bot.send_message(
        callback_query.from_user.id,
        'Введите сумму, которую хотите вычесть:'
    )


@user_dp.message_handler(state=AdminSubtractBalance.balance)
@allow_bace_access
async def admin_subtract_user_balance_state(msg: types.Message, state: FSMContext):
    async with state.proxy() as data:
        data["balance"] = msg.text
    await state.finish()
    await subtract_balance(int(data['user_id']), int(data['balance']))
    await user_bot.send_message(
        msg.from_user.id,
        'Действие выполнено успешно!'
    )

##

## Задания

@user_dp.callback_query_handler(lambda c: 'admin_show_user_tasks' in c.data)
@allow_bace_access
async def admin_show_user_task(callback_query: types.CallbackQuery):
    user_id = callback_query.data.split(':')[1]
    await user_bot.send_message(
        callback_query.from_user.id,
        'Выберите:',
        reply_markup=admin_show_user_task_kb(user_id)
    )


# Зачесть выполнение
@user_dp.callback_query_handler(lambda c: 'admin_give_done_task' in c.data)
@allow_bace_access
async def admin_give_done_task(callback_query: types.CallbackQuery, state: FSMContext):
    user_id = callback_query.data.split(':')[1]
    await AdminGiveTask.number_task.set()
    async with state.proxy() as data:
        data["user_id"] = user_id
    await AdminGiveTask.number_task.set()
    await user_bot.send_message(
        callback_query.from_user.id,
        'Введите номер задания'
    )

@user_dp.message_handler(state=AdminGiveTask.number_task)
@allow_bace_access
async def admin_give_done_task_state(msg: types.Message, state: FSMContext):
    async with state.proxy() as data:
        data["number_task"] = msg.text
    await state.finish()
    message = await user_bot.send_message(
        msg.from_user.id,
        'Немного подождите, засчитываем пользователю задачу...'
    )
    if int(data['number_task']) in tasks_progress:
        if int(data["user_id"]) in tasks_progress[int(data["number_task"])]['users']['in_process']:
            del tasks_progress[int(data["number_task"])]['users']['in_process'][int(data["user_id"])]
        if int(data["user_id"]) in tasks_progress[int(data["number_task"])]['users']['rejected']:
            del tasks_progress[int(data["number_task"])]['users']['rejected'][int(data["user_id"])]
        if int(data["user_id"]) in tasks_progress[int(data["number_task"])]['users']['canceled']:
            tasks_progress[int(data["number_task"])]['users']['canceled'].remove(int(data["user_id"]))
        if int(data["user_id"]) in tasks_progress[int(data['number_task'])]['users']['checking']:
            del tasks_progress[int(data['number_task'])]['users']['checking'][int(data["user_id"])]
        if int(data['user_id']) not in tasks_progress[int(data['number_task'])]['users']['done']:
            tasks_progress[int(data['number_task'])]['users']['done'][int(data['user_id'])] = {
                'start_date': '',
                'end_date': datetime.datetime.now(),
                'admin_id': msg.from_user.id,
                'admin_name': msg.from_user.username
            }
            task_datas = await get_task_datas(int(data["number_task"]))
            await adding_stat_accepted()
            user_datas = await get_user_data(data["user_id"])
            if user_datas["who_invite"] != '':
                await adding_ref_balance(user_datas["who_invite"], int(int(task_datas["price"]) * 0.15))
            await adding_balance(data["user_id"], int(task_datas["price"]))
            if data['number_task'] in await get_all_active_tasks(data["user_id"]):
                await delete_active_task(data["user_id"], data['number_task'])
            if data['number_task'] not in await get_all_history_tasks(data["user_id"]):
                await adding_new_history_task(data["user_id"], data["number_task"])
            if data['number_task'] not in await get_all_done_tasks(data["user_id"]):
                await adding_new_done_task(data["user_id"], data['number_task'])
            await message.delete()
            await user_bot.send_message(
                msg.from_user.id,
                'Успешно!'
            )
        else:
            await message.delete()
            await user_bot.send_message(
                msg.from_user.id,
                'Задача и так засчитана пользователю'
            )
    elif int(data['number_task']) in await get_all_archive_tasks_list():
        if int(data["user_id"]) not in archive_tasks_progerss[int(data['number_task'])]['users']['done']:
            if int(data["user_id"]) in archive_tasks_progerss[int(data["number_task"])]['users']['in_process']:
                del archive_tasks_progerss[int(data["number_task"])]['users']['in_process'][int(data["user_id"])]
            if int(data["user_id"]) in archive_tasks_progerss[int(data["number_task"])]['users']['rejected']:
                del archive_tasks_progerss[int(data["number_task"])]['users']['rejected'][int(data["user_id"])]
            if int(data["user_id"]) in archive_tasks_progerss[int(data["number_task"])]['users']['canceled']:
                archive_tasks_progerss[int(data["number_task"])]['users']['canceled'].remove(int(data["user_id"]))
            if int(data["user_id"]) in archive_tasks_progerss[int(data['number_task'])]['users']['checking']:
                del archive_tasks_progerss[int(data['number_task'])]['users']['checking'][int(data["user_id"])]
            archive_tasks_progerss[int(data['number_task'])]['users']['done'][int(data["user_id"])] = {
                'start_date': '',
                'end_date': datetime.datetime.now(),
                'admin_id': msg.from_user.id,
                'admin_name': msg.from_user.username
            }
            if data['number_task'] in await get_all_active_tasks(data["user_id"]):
                await delete_active_task(data["user_id"], data['number_task'])
            if data['number_task'] not in await get_all_history_tasks(data["user_id"]):
                await adding_new_history_task(data["user_id"], data["number_task"])
            if data['number_task'] not in await get_all_done_tasks(data["user_id"]):
                await adding_new_done_task(data["user_id"], data['number_task'])
            task_datas = await get_archive_task_datas(int(data['number_task']))
            await adding_balance(int(data["user_id"]), int(task_datas["price"]))
            user_datas = await get_user_data(data["user_id"])
            if user_datas["who_invite"] != '':
                await adding_ref_balance(user_datas["who_invite"], int(int(task_datas["price"]) * 0.15))
            await adding_stat_accepted()
            await message.delete()
            await user_bot.send_message(
                msg.from_user.id,
                'Успешно!'
            )
        else:
            await message.delete()
            await user_bot.send_message(
                msg.from_user.id,
                'Задача и так засчитана пользователю'
            )

    else:
        await message.delete()
        await user_bot.send_message(
            msg.from_user.id,
            'Похоже задачи с таким номером не существует, проверьть всё еще раз!'
        )

#

# Все задания
@user_dp.callback_query_handler(lambda c: 'admin_show_all_user_task' in c.data)
@allow_bace_access
async def admin_show_all_user_task(callback_query: types.CallbackQuery, state: FSMContext):
    user_id = callback_query.data.split(':')[1]
    await AdminSearchUserTask.user_id.set()
    async with state.proxy() as data:
        data['user_id'] = user_id
    await AdminSearchUserTask.number_task.set()
    message = await user_bot.send_message(
        callback_query.from_user.id,
        'Немного подождите, получаем статистику польлователя...'
    )
    done_tasks = []
    done_tasks_with_price = []
    for task in tasks_progress:
        if int(user_id) in tasks_progress[task]['users']['done']:
            done_tasks.append(task)
    for task in archive_tasks_progerss:
        if int(user_id) in archive_tasks_progerss[task]['users']['done']:
            done_tasks.append(task) 
    if len(done_tasks) != 0:
        for task in done_tasks:
            task_datas = await get_task_datas(int(task))
            if task_datas:
                done_tasks_with_price.append((task, task_datas["price"]))
        workbook = openpyxl.Workbook()
        
        # Выбираем активный лист
        sheet = workbook.active
        
        # Заголовки столбцов
        sheet['A1'] = 'Задание'
        sheet['B1'] = 'Цена'

        for index, (user_id, name) in enumerate(done_tasks_with_price, start=2):
            sheet.cell(row=index, column=1, value=user_id)
            sheet.cell(row=index, column=2, value=name)
        
        # Сохраняем файл
        workbook.save(f'static/excel/{data["user_id"]}.xlsx')
        await user_bot.send_document(
            callback_query.from_user.id,
            document=open(f'static/excel/{data["user_id"]}.xlsx', 'rb')
        )
        

    await user_bot.send_message(
        callback_query.from_user.id,
        'Выберите или введите номер задания',
        reply_markup=await admin_show_all_user_task_kb(page=0, user_id=data['user_id'], tasks_progerss=tasks_progress, archive_task_progress=archive_tasks_progerss)
    )
    await message.delete()

@user_dp.callback_query_handler(lambda c: 'admin_show_last_all_user_task' in c.data, state=AdminSearchUserTask.number_task)
@allow_bace_access
async def admin_show_last_all_user_task_last(callback_query: types.CallbackQuery, state: FSMContext):
    page = int(callback_query.data.split(':')[1])
    user_id = int(callback_query.data.split(':')[2])
    await callback_query.message.edit_text(
        text='Выберите или введите номер задания',
        reply_markup=await admin_show_all_user_task_kb(page=page, user_id=user_id, tasks_progerss=tasks_progress, archive_task_progress=archive_tasks_progerss)
    )
    return

@user_dp.callback_query_handler(lambda c: 'admin_show_next_all_user_task' in c.data, state=AdminSearchUserTask.number_task)
@allow_bace_access
async def admin_show_next_all_user_task_last(callback_query: types.CallbackQuery, state: FSMContext):
    page = int(callback_query.data.split(':')[1])
    user_id = int(callback_query.data.split(':')[2])
    await callback_query.message.edit_text(
        text='Выберите или введите номер задания',
        reply_markup=await admin_show_all_user_task_kb(page=page, user_id=user_id, tasks_progerss=tasks_progress, archive_task_progress=archive_tasks_progerss)
    )
    return

@user_dp.callback_query_handler(lambda c: 'admin_show_all_user_task' in c.data, state=AdminSearchUserTask.number_task)
@allow_bace_access
async def admin_show_all_user_task_state(callback_query: types.CallbackQuery, state: FSMContext):
    await state.finish()
    user_id = int(callback_query.data.split(':')[1])
    number_task = callback_query.data.split(':')[2]
    await admin_show_all_user_task(callback_query, user_id, number_task)

@user_dp.message_handler(state=AdminSearchUserTask.number_task)
@allow_bace_access
async def admin_show_all_user_task_state_step(msg: types.Message, state: FSMContext):
    async with state.proxy() as data:
        data["number_task"] = msg.text
    await state.finish()
    await admin_show_all_user_task(msg, data["user_id"], data['number_task'])



@allow_bace_access
async def admin_show_all_user_task(msg: types.Message, user_id, number_task):
    message = await user_bot.send_message(
        msg.from_user.id,
        'Немного подождите, собираем ифнормацию по пользваотелю...'
    )
    try:
        task_datas = await get_task_datas(int(number_task))
        if task_datas:
            if 'Задание перенесено в архив' in task_datas["full_text"]:
                if int(user_id) in archive_tasks_progerss[int(number_task)]['users']['done']:
                    if archive_tasks_progerss[int(number_task)]['users']['done'][int(user_id)]['start_date'] == '':
                        end_date = archive_tasks_progerss[int(number_task)]['users']['done'][int(user_id)]['end_date']
                        admin_id = archive_tasks_progerss[int(number_task)]['users']['done'][int(user_id)]['admin_id']
                        admin_name = archive_tasks_progerss[int(number_task)]['users']['done'][int(user_id)]['admin_name']
                        await message.delete()
                        await user_bot.send_message(
                            msg.from_user.id,
                            f'Задание зачтено вручную\n\nВремя и дата: {end_date}\nId админа: {admin_id}\n\nUsername админа: {admin_name}'
                        )
                    else:
                        start_date = archive_tasks_progerss[int(number_task)]['users']['done'][int(user_id)]['start_date']
                        end_date = archive_tasks_progerss[int(number_task)]['users']['done'][int(user_id)]['end_date']
                        admin_id = archive_tasks_progerss[int(number_task)]['users']['done'][int(user_id)]['admin_id']
                        admin_name = archive_tasks_progerss[int(number_task)]['users']['done'][int(user_id)]['admin_name']
                        await message.delete()
                        await user_bot.send_message(
                            msg.from_user.id,
                            f'Задание зачтено через выполнения\n\nВремя начала выполнения: {start_date}\nВремя сдачи задания: {end_date}\nId админа: {admin_id}\n\nUsername админа: {admin_name}'
                        )
                        all_items = os.listdir(f'static/archive/{number_task}')
                        files = [item for item in all_items if os.path.isfile(os.path.join(f'static/archive/{number_task}', item))]
                        media = []
                        for file in files:
                            if str(user_id) in file:
                                if 'mp4' in file:
                                    media.append(types.InputMediaVideo(media=open(f'static/archive/{number_task}/{file}', 'rb')))
                                elif 'jpg' in file:
                                    media.append(types.InputMediaPhoto(media=open(f'static/archive/{number_task}/{file}', 'rb')))
                        
                        if media:
                            await user_bot.send_media_group(
                                msg.from_user.id, 
                                media=media,
                            )
                elif int(user_id) in archive_tasks_progerss[int(number_task)]['users']['rejected']:
                    reason = archive_tasks_progerss[int(number_task)]['users']['rejected'][int(user_id)]['reason']
                    start_date = archive_tasks_progerss[int(number_task)]['users']['rejected'][int(user_id)]['start_date']
                    end_date = archive_tasks_progerss[int(number_task)]['users']['rejected'][int(user_id)]['end_date']
                    admin_id = archive_tasks_progerss[int(number_task)]['users']['rejected'][int(user_id)]['admin_id']
                    admin_name = archive_tasks_progerss[int(number_task)]['users']['rejected'][int(user_id)]['admin_name']
                    await message.delete()
                    await user_bot.send_message(
                        msg.from_user.id,
                        f'Задание не принято\n\nВремя начала выполнения: {start_date}\nВремя сдачи задания: {end_date}\nId админа: {admin_id}\n\nUsername админа: {admin_name}\n\nПричина отклонения: {reason}'
                    )
                    all_items = os.listdir(f'static/archive/{number_task}')
                    files = [item for item in all_items if os.path.isfile(os.path.join(f'static/archive/{number_task}', item))]
                    media = []
                    for file in files:
                        if str(user_id) in file:
                            if 'mp4' in file:
                                media.append(types.InputMediaVideo(media=open(f'static/archive/{number_task}/{file}', 'rb')))
                            elif 'jpg' in file:
                                media.append(types.InputMediaPhoto(media=open(f'static/archive/{number_task}/{file}', 'rb')))
                        
                    if media:
                        await user_bot.send_media_group(
                            msg.from_user.id, 
                            media=media,
                        )
            else:
                if int(user_id) in tasks_progress[int(number_task)]['users']['done']:
                    if tasks_progress[int(number_task)]['users']['done'][int(user_id)]['start_date'] == '':
                        end_date = tasks_progress[int(number_task)]['users']['done'][int(user_id)]['end_date']
                        admin_id = tasks_progress[int(number_task)]['users']['done'][int(user_id)]['admin_id']
                        admin_name = tasks_progress[int(number_task)]['users']['done'][int(user_id)]['admin_name']
                        await message.delete()
                        await user_bot.send_message(
                            msg.from_user.id,
                            f'Задание зачтено вручную\n\nВремя и дата: {end_date}\nId админа: {admin_id}\n\nUsername админа: {admin_name}'
                        )
                    else:
                        start_date = tasks_progress[int(number_task)]['users']['done'][int(user_id)]['start_date']
                        end_date = tasks_progress[int(number_task)]['users']['done'][int(user_id)]['end_date']
                        admin_id = tasks_progress[int(number_task)]['users']['done'][int(user_id)]['admin_id']
                        admin_name = tasks_progress[int(number_task)]['users']['done'][int(user_id)]['admin_name']
                        await message.delete()
                        await user_bot.send_message(
                            msg.from_user.id,
                            f'Задание зачтено через выполнения\n\nВремя начала выполнения: {start_date}\nВремя сдачи задания: {end_date}\nId админа: {admin_id}\n\nUsername админа: {admin_name}'
                        )
                        all_items = os.listdir(f'static/archive/{number_task}')
                        files = [item for item in all_items if os.path.isfile(os.path.join(f'static/archive/{number_task}', item))]
                        media = []
                        for file in files:
                            if str(user_id) in file:
                                if 'mp4' in file:
                                    media.append(types.InputMediaVideo(media=open(f'static/archive/{number_task}/{file}', 'rb')))
                                elif 'jpg' in file:
                                    media.append(types.InputMediaPhoto(media=open(f'static/archive/{number_task}/{file}', 'rb')))
                        
                        if media:
                            await user_bot.send_media_group(
                                msg.from_user.id, 
                                media=media,
                            )
                elif int(user_id) in tasks_progress[int(number_task)]['users']['rejected']:
                    reason = tasks_progress[int(number_task)]['users']['rejected'][int(user_id)]['reason']
                    start_date = tasks_progress[int(number_task)]['users']['rejected'][int(user_id)]['start_date']
                    end_date = tasks_progress[int(number_task)]['users']['rejected'][int(user_id)]['end_date']
                    admin_id = tasks_progress[int(number_task)]['users']['rejected'][int(user_id)]['admin_id']
                    admin_name = tasks_progress[int(number_task)]['users']['rejected'][int(user_id)]['admin_name']
                    await message.delete()
                    await user_bot.send_message(
                        msg.from_user.id,
                        f'Задание не принято\n\nВремя начала выполнения: {start_date}\nВремя сдачи задания: {end_date}\nId админа: {admin_id}\n\nUsername админа: {admin_name}\n\nПричина отклонения: {reason}'
                    )
                    all_items = os.listdir(f'static/archive/{number_task}')
                    files = [item for item in all_items if os.path.isfile(os.path.join(f'static/archive/{number_task}', item))]
                    media = []
                    for file in files:
                        if str(user_id) in file:
                            if 'mp4' in file:
                                media.append(types.InputMediaVideo(media=open(f'static/archive/{number_task}/{file}', 'rb')))
                            elif 'jpg' in file:
                                media.append(types.InputMediaPhoto(media=open(f'static/archive/{number_task}/{file}', 'rb')))
                    
                    if media:
                        await user_bot.send_media_group(
                            msg.from_user.id, 
                            media=media,
                        )
        else:
            await message.delete()
            await user_bot.send_message(
                msg.from_user.id,
                'Похоже такого задания не существует'
            )
    except:
        await message.delete()
        await user_bot.send_message(
            msg.from_user.id,
            'Данные по заданию не найдены...'
        )

##


#######


####### Общая статистика

@user_dp.callback_query_handler(lambda c: c.data == 'admin_stat')
@allow_vip_access
async def admin_stat(msg: types.Message):
    stat_datas = await get_stat()
    if stat_datas:
        count_users = await get_users_count()
        await user_bot.send_message(
            msg.from_user.id,
            f'Денег выплачено: {stat_datas["paid"]}\nОсталось выплатить: \nВсего человек: {count_users}\nВсего заданий: {stat_datas["count_tasks"]}\nАктивных: {int(stat_datas["count_tasks"]) - int(stat_datas["delete_tasks"])}\nУдаленных: {stat_datas["delete_tasks"]}\nВыполненных:\nВыполнений принято: {stat_datas["accepted"]}\nВыполнений отклонено: {stat_datas["rejected"]}',
            reply_markup=admin_user_stat_kb()
        )



## Выплаты

@user_dp.callback_query_handler(lambda c: c.data == 'admip_paid')
@allow_vip_access
async def admin_paid(callback_query: types.CallbackQuery):
    await user_bot.send_message(
        callback_query.from_user.id,
        'Выберите',
        reply_markup=admin_user_paid_kb()
    )

@user_dp.callback_query_handler(lambda c: c.data == 'current_transactions')
@allow_vip_access
async def admin_all_transactions(callback_query: types.CallbackQuery):
    if os.path.exists('static/transactions/list.xlsx'):
        await user_bot.send_document(
            callback_query.from_user.id,
            document=open('static/transactions/list.xlsx', 'rb')
        )
        await user_bot.send_message(
            callback_query.from_user.id,
            'Обнулить список?',
            reply_markup=admin_update_user_paid_kb()
        )
    else:
        users = await get_users_for_paid()
        if users:
            workbook = openpyxl.Workbook()
        
            # Выбираем активный лист
            sheet = workbook.active
            
            # Заголовки столбцов
            sheet['A1'] = 'user_id'
            sheet['B1'] = 'username'
            sheet['C1'] = 'баланс'
            sheet['D1'] = 'Тип банка'
            sheet['E1'] = 'Номер карты'
            sheet['F1'] = 'Имя банка'
            sheet['G1'] = 'Номер телефона'
            for index, (user_id, name, balance, type_bank, card_number, bank_name, phone_number) in enumerate(users, start=2):
                sheet.cell(row=index, column=1, value=user_id)
                sheet.cell(row=index, column=2, value=name)
                sheet.cell(row=index, column=3, value=balance)
                sheet.cell(row=index, column=4, value=type_bank)
                sheet.cell(row=index, column=5, value=card_number)
                sheet.cell(row=index, column=6, value=bank_name)
                sheet.cell(row=index, column=7, value=phone_number)
            workbook.save('static/transactions/list.xlsx')
            await user_bot.send_document(
                callback_query.from_user.id,
                document=open('static/transactions/list.xlsx', 'rb')
            )
            await user_bot.send_message(
                callback_query.from_user.id,
                'Обнулить список?',
                reply_markup=admin_update_user_paid_kb()
            )
        else:
            await user_bot.send_message(
                callback_query.from_user.id,
                'Похоже нет пользователей с балансом больше 200'
            )

@user_dp.callback_query_handler(lambda c: c.data == 'admin_back_update_paid')
@allow_vip_access
async def admin_update_paid_back(callback_query: types.CallbackQuery):
    await admin_cmd_start(callback_query)

@user_dp.callback_query_handler(lambda c: c.data == 'admin_update_paid_list')
@allow_vip_access
async def admin_update_paid(callback_query: types.CallbackQuery):
    if os.path.exists('static/transactions/list.xlsx'):
        message = await user_bot.send_message(
            callback_query.from_user.id,
            'Подождите, производится снятие баланса...'
        )
        users = []
        workbook = openpyxl.load_workbook('static/transactions/list.xlsx')
        # Выбираем активный лист
        sheet = workbook.active

        # Создаем пустой список для хранения данных
        users = []

        # Проходим по строкам таблицы, начиная со второй строки (первая строка - заголовки)
        for row in sheet.iter_rows(min_row=2, values_only=True):
            user_id, name, balance, type_bank, card_number, bank_name, phone_number = row
            users.append((user_id, name, balance, type_bank, card_number, bank_name, phone_number))
        for (user_id, name, balance, type_bank, card_number, bank_name, phone_number) in users:
            try:
                await subtract_balance(user_id, balance)
                await adding_stat_paid(balance)
                await create_transaction(
                    user_id=user_id,
                    user_name=name,
                    amount=int(balance),
                    date=datetime.datetime.now()
                )
            except:
                pass
        os.remove('static/transactions/list.xlsx')
        await message.delete()
        await user_bot.send_message(
            callback_query.from_user.id,
            'Успешно!'
        )
    else:
        await user_bot.send_message(
            callback_query.from_user.id,
            'Что-то пошло не так...'
        )


@user_dp.callback_query_handler(lambda c: c.data == 'all_transactions')
@allow_vip_access
async def admin_all_transactions(callback_query: types.CallbackQuery):
    transactions = await get_all_transactions()
    if transactions:
        message = await user_bot.send_message(
            callback_query.from_user.id,
            'Немного подождите, формируется таблица...'
        )
        workbook = openpyxl.Workbook()
    
        # Выбираем активный лист
        sheet = workbook.active
            
        # Заголовки столбцов
        sheet['A1'] = 'user_id'
        sheet['B1'] = 'username'
        sheet['C1'] = 'Сумма'
        sheet['D1'] = 'Дата и время'

        for index, (user_id, user_name, amount, date) in enumerate(transactions, start=2):
            sheet.cell(row=index, column=1, value=user_id)
            sheet.cell(row=index, column=2, value=user_name)
            sheet.cell(row=index, column=3, value=amount)
            sheet.cell(row=index, column=4, value=str(date))
        workbook.save('static/transactions/history.xlsx')
        await message.delete()
        await user_bot.send_document(
            callback_query.from_user.id,
            document=open('static/transactions/history.xlsx', 'rb')
        )
    else:
        await user_bot.send_message(
            callback_query.from_user.id,
            'Нет никаких записей...'
        )




##

## Рефералы

@user_dp.callback_query_handler(lambda c: c.data == 'admin_ref')
@allow_vip_access
async def admin_ref(callback_query: types.CallbackQuery):
    await AdminSearchRef.user_id.set()
    await user_bot.send_message(
        callback_query.from_user.id,
        'Введите id пользователя'
    )

@user_dp.message_handler(state=AdminSearchRef.user_id)
@allow_vip_access
async def admin_ref_state(msg: types.Message, state: FSMContext):
    async with state.proxy() as data:
        data["user_id"] = msg.text
    await state.finish()
    users = await get_all_ref(data["user_id"])
    if users:
        await user_bot.send_message(
            msg.from_user.id,
            f'Количество приглашений: {len(users)}\n\nСписок проглашенных пользователей: {" ".join(users)}'
        )
    else:
        await user_bot.send_message(
            msg.from_user.id,
            'Похоже такого пользоватлея не существует'
        )

##

## Топ активности

@user_dp.callback_query_handler(lambda c: c.data == 'admin_top')
@allow_vip_access
async def admin_top_activity(callback_query: types.CallbackQuery):
    message = await user_bot.send_message(
        callback_query.from_user.id,
        'Немного подождите, формируется топ активности...'
    )
    top_list = await get_top_list()
    top = ''
    for index, (user_id, activity) in enumerate(top_list, start=1):
        top += f'{index}. {user_id} - {activity} выполненных заданий\n'
    await message.delete()
    await user_bot.send_message(
        callback_query.from_user.id,
        f'Топ 10 пользователей по активности:\n\n{top}'
    )

##

## Предупреждения и блокировки

@user_dp.callback_query_handler(lambda c: c.data == 'admin_warn_and_block')
@allow_vip_access
async def admin_warn_and_blocks(callback_query: types.CallbackQuery):
    message = await user_bot.send_message(
        callback_query.from_user.id,
        'Немного подождите, формируются таблицы...'
    )
    workbook = openpyxl.Workbook()
    
    # Выбираем активный лист
    sheet = workbook.active
            
    # Заголовки столбцов
    sheet['A1'] = 'user_id'
    for index, user_id in enumerate(block_users, start=2):
        sheet.cell(row=index, column=1, value=user_id)
    workbook.save('static/excel/block_users.xlsx')
    workbook = openpyxl.Workbook()
    
    # Выбираем активный лист
    sheet = workbook.active
            
    # Заголовки столбцов
    sheet['A1'] = 'user_id'
    sheet['B1'] = 'количество предупреждений'
    for index, user_id in enumerate(warnings, start=2):
        sheet.cell(row=index, column=1, value=user_id)
        sheet.cell(row=index, column=2, value=len(warnings[int(user_id)]))
    workbook.save('static/excel/warnings.xlsx')
    await message.delete()
    await user_bot.send_document(
        callback_query.from_user.id,
        document=open('static/excel/warnings.xlsx', 'rb'),
        caption='Список предупреждений',
        reply_markup=admin_search_warn()

    )
    await user_bot.send_document(
        callback_query.from_user.id,
        document=open('static/excel/block_users.xlsx', 'rb'),
        caption='Список заблокированных пользователей',
        reply_markup=admin_unblock_user_kb()
    )

@user_dp.callback_query_handler(lambda c: c.data == 'admin_unblock_user')
@allow_vip_access
async def admin_unblock_user(callback_query: types.CallbackQuery):
    await AdminUnblockUser.user_id.set()
    await user_bot.send_message(
        callback_query.from_user.id,
        'Введите id пользователя, которого хотите разблокировать:'
    )

@user_dp.message_handler(state=AdminUnblockUser.user_id)
@allow_vip_access
async def admin_unblock_user_state(msg: types.Message, state: FSMContext):
    async with state.proxy() as data:
        data["user_id"] = msg.text
    await state.finish()
    if await unblock_user(data['user_id']):
        if int(data['user_id']) in block_users:
            block_users.remove(int(data['user_id']))
        await user_bot.send_message(
            msg.from_user.id,
            'Пользователь разблокирован!'
        )
    else:
        await user_bot.send_message(
            msg.from_user.id,
            'Похоже такого пользователя не существует'
        )

@user_dp.callback_query_handler(lambda c: c.data == 'admin_search_warn')
@allow_vip_access
async def admin_searc_warn(callback_query: types.CallbackQuery):
    await AdminSearchWarn.user_id.set()
    await user_bot.send_message(
        callback_query.from_user.id,
        'Введите id пользователя'
    )

@user_dp.message_handler(state=AdminSearchWarn.user_id)
@allow_vip_access
async def admin_searc_warn_state(msg: types.Message, state: FSMContext):
    async with state.proxy() as data:
        data["user_id"] = msg.text
    await state.finish()
    warning = ''
    if int(data["user_id"]) in warnings:
        for index, war in enumerate(warnings[int(data["user_id"])], start=1):
            warning += f'{index}: {war}\n'
        warn_len = len(warnings[int(data["user_id"])])
    if len(warning) != 0:
        await user_bot.send_message(
            msg.from_user.id,
            f'У пользователя уже {warn_len} предупреждений.\n\n{warning}',
            reply_markup=admin_users_warnings_kb(data["user_id"])
        )
    else:
        await user_bot.send_message(
            msg.from_user.id,
            'У данного пользователя нет предупреждений, либо вы ошиблись в ID'
        )

##

    
####### Рассылка

@user_dp.callback_query_handler(lambda c: c.data == 'admin_message')
@allow_vip_access
async def admin_message(callback_query: types.CallbackQuery):
    await AdminMakeMessage.text.set()
    await user_bot.send_message(
        callback_query.from_user.id,
        'Введите текст рассылки',
    )

@user_dp.message_handler(state=AdminMakeMessage.text)
@allow_vip_access
async def admin_message_state(msg: types.Message, state: FSMContext):
    async with state.proxy() as data:
        data["text"] = msg.text
    await AdminMakeMessage.conf.set()
    await user_bot.send_message(
        msg.from_user.id,
        'Сообщение будет выглядеть вот так:\n\nВы уверены?'
    )
    await user_bot.send_message(
        msg.from_user.id,
        data['text'],
        reply_markup=admin_message_conf()
    )

@user_dp.callback_query_handler(lambda c: c.data == 'admin_make_message_conf', state=AdminMakeMessage.conf)
@allow_vip_access
async def admin_message_state_conf(callback_query: types.CallbackQuery, state: FSMContext):
    async with state.proxy() as data:
        data["conf"] = 'Yes'
    await state.finish()
    message = await user_bot.send_message(
        callback_query.from_user.id,
        'Подождите. идет рассылка...'
    )
    users = await get_user_list_for_message()
    if users:
        async def send_message_to_user(user_id):
            try:
                await user_bot.send_message(
                    int(user_id),
                    data["text"],
                    parse_mode='HTML',
                    reply_markup=user_edit_text_kb(),
                )
                return True
            except:
                return False

        tasks = [send_message_to_user(user) for user in users]
        results = await asyncio.gather(*tasks)
        await message.delete()
        await user_bot.send_message(
            callback_query.from_user.id,
            f'Сообщение отправлено {results} пользователям'
        )

@user_dp.callback_query_handler(lambda c: c.data == 'admin_make_message_cancel', state=AdminMakeMessage.conf)
@allow_vip_access
async def admin_message_state_cancel(callback_query: types.CallbackQuery, state: FSMContext):
    await state.finish()
    await user_bot.send_message(
        callback_query.from_user.id,
        'Действие отменено'
    )


#######

from utils.dump import dump_dicts

@user_dp.message_handler(commands='create_backups')
@allow_vip_access
async def admin_create_backups(msg: types.Message):
    up = await dump_dicts(cached_data, main_menu_icon, tasks_progress, archive_tasks_progerss, warnings)
    if up:
        await user_bot.send_message(
            msg.from_user.id,
            'Бэкап создан'
        )
    else:
        await user_bot.send_message(
            msg.from_user.id,
            'Что-то пошло не так'
        )



####### Подгрузка бэкапов



def load_data():
    global cached_data, main_menu_icon, tasks_progress
    if os.path.exists('backups/cached_data.pkl'):
        with open('backups/cached_data.pkl', 'rb') as file:
            cached_data = pickle.load(file)
    if os.path.exists('backups/main_menu_icon.pkl'):
        with open('backups/main_menu_icon.pkl', 'rb') as file:
            main_menu_icon = pickle.load(file)
    if os.path.exists('backups/tasks_progress.pkl'):
        with open('backups/tasks_progress.pkl', 'rb') as file:
            tasks_progress = pickle.load(file)
    if os.path.exists('backups/archive_tasks_progerss.pkl'):
        with open('backups/archive_tasks_progerss.pkl', 'rb') as file:
            archive_tasks_progerss = pickle.load(file)
    if os.path.exists('backups/warnings.pkl'):
        with open('backups/warnings.pkl', 'rb') as file:
            warnings = pickle.load(file)




#######


if __name__ == '__main__':
    start_db()
    load_data()
    create_statistic()
    block_users = get_list_of_blocks()
    from utils.scheduler import start_schedule

    scheduler = AsyncIOScheduler()
    thread_backup_dicts = threading.Thread(target=start_schedule, daemon=True, args=(scheduler, cached_data, main_menu_icon, tasks_progress, archive_tasks_progerss, warnings))
    thread_backup_dicts.start()
    scheduler.start()
    
    executor.start_polling(user_dp, skip_updates=True)
